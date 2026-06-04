#!/usr/bin/env python3
"""项目管理系统 Flask 服务"""

from flask import Flask, jsonify, request, send_file
import sqlite3
from datetime import datetime, timedelta
import subprocess
import os
import hashlib
import secrets
import re
import json
from functools import wraps

# 前端文件目录
FRONTEND_DIR = os.environ.get('PM_FRONTEND_DIR', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'frontend'))

app = Flask(__name__, static_folder=FRONTEND_DIR)
DB_PATH = os.environ.get('PM_DB_PATH', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'pm.db'))
MEMOS_SCRIPT = os.environ.get('MEMOS_SCRIPT', '')

# 用户会话存储（简单实现）
USER_SESSIONS = {}
SESSION_EXPIRE_HOURS = 24

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# ========== 用户认证系统 ==========

def init_user_table():
    """初始化用户表"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS pm_user (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            person_id INTEGER,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT DEFAULT 'user',
            last_login TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (person_id) REFERENCES person(id)
        )
    ''')
    # 项目查看者授权表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS project_viewer (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            person_id INTEGER NOT NULL,
            granted_by INTEGER,
            granted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(project_id, person_id),
            FOREIGN KEY (project_id) REFERENCES project(id),
            FOREIGN KEY (person_id) REFERENCES person(id)
        )
    ''')
    conn.commit()
    conn.close()

def hash_password(password: str) -> str:
    """密码哈希"""
    return hashlib.sha256(password.encode()).hexdigest()

def generate_username(name: str, existing_names: set) -> str:
    """生成用户名（姓名拼音缩写）"""
    from pypinyin import lazy_pinyin, Style
    # 获取姓名拼音首字母
    pinyin_list = lazy_pinyin(name, style=Style.FIRST_LETTER)
    base_name = ''.join(pinyin_list).lower()
    
    # 检查是否重复
    if base_name not in existing_names:
        return base_name
    
    # 重复则加数字序号
    counter = 1
    while f"{base_name}{counter:02d}" in existing_names:
        counter += 1
    return f"{base_name}{counter:02d}"

def create_users_from_persons():
    """根据person表创建用户账号"""
    init_user_table()
    conn = get_db()
    cursor = conn.cursor()
    
    # 获取已存在的用户名
    cursor.execute('SELECT username FROM pm_user')
    existing_names = set(row['username'] for row in cursor.fetchall())
    
    # 获取所有person
    cursor.execute('SELECT id, name FROM person WHERE status = "active"')
    persons = cursor.fetchall()
    
    # 默认密码
    default_password = "pm2026"
    password_hash = hash_password(default_password)
    
    created_count = 0
    for person in persons:
        username = generate_username(person['name'], existing_names)
        try:
            cursor.execute('''
                INSERT INTO pm_user (person_id, username, password_hash, role)
                VALUES (?, ?, ?, 'user')
            ''', (person['id'], username, password_hash))
            existing_names.add(username)
            created_count += 1
            print(f"创建账号: {username} -> {person['name']}")
        except sqlite3.IntegrityError:
            pass  # 用户已存在
    
    conn.commit()
    conn.close()
    return created_count

def generate_token():
    """生成会话token"""
    return secrets.token_hex(32)

def check_auth(f):
    """认证装饰器"""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '')
        if token.startswith('Bearer '):
            token = token[7:]
        
        if not token or token not in USER_SESSIONS:
            return jsonify({'success': False, 'error': '未登录或会话已过期'}), 401
        
        session = USER_SESSIONS[token]
        if datetime.now() > session['expire_time']:
            del USER_SESSIONS[token]
            return jsonify({'success': False, 'error': '会话已过期'}), 401
        
        request.current_user = session['user']
        return f(*args, **kwargs)
    return decorated

def require_admin(f):
    """管理员权限装饰器"""
    @wraps(f)
    def decorated(*args, **kwargs):
        user = getattr(request, 'current_user', None)
        if not user:
            return jsonify({'success': False, 'error': '未登录'}), 401
        if user.get('role') != 'admin':
            return jsonify({'success': False, 'error': '需要管理员权限'}), 403
        return f(*args, **kwargs)
    return decorated

def is_project_owner(project_id, person_id):
    """检查是否是项目负责人"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT owner_id FROM project WHERE id = ?', (project_id,))
    row = cursor.fetchone()
    conn.close()
    return row and row['owner_id'] == person_id

def is_project_participant(project_id, person_id):
    """检查是否是项目参与者（owner或该项目有任务分配给该用户）"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT 1 FROM project p
        LEFT JOIN task t ON t.project_id = p.id
        WHERE p.id = ? AND (p.owner_id = ? OR t.assignee_id = ?)
        LIMIT 1
    ''', (project_id, person_id, person_id))
    result = cursor.fetchone()
    conn.close()
    return result is not None

def is_project_viewer(project_id, person_id):
    """检查是否是项目查看者（被授权查看该项目的用户）"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT 1 FROM project_viewer WHERE project_id = ? AND person_id = ?', (project_id, person_id))
    result = cursor.fetchone()
    conn.close()
    return result is not None

def get_user_project_ids(person_id, is_admin=False):
    """获取用户可见的所有项目ID列表（参与+被授权查看的）"""
    if is_admin:
        return None  # None表示不过滤
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT DISTINCT p.id FROM project p
        LEFT JOIN task t ON t.project_id = p.id
        LEFT JOIN project_viewer pv ON pv.project_id = p.id
        WHERE p.owner_id = ? OR t.assignee_id = ? OR pv.person_id = ?
    ''', (person_id, person_id, person_id))
    result = [row['id'] for row in cursor.fetchall()]
    conn.close()
    return result

@app.route('/api/auth/login', methods=['POST'])
def login():
    """用户登录"""
    data = request.json
    username = data.get('username', '').lower()
    password = data.get('password', '')
    
    if not username or not password:
        return jsonify({'success': False, 'error': '请输入用户名和密码'})
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT u.*, p.name as person_name
        FROM pm_user u
        LEFT JOIN person p ON u.person_id = p.id
        WHERE u.username = ?
    ''', (username,))
    user = cursor.fetchone()
    
    if not user:
        return jsonify({'success': False, 'error': '用户名不存在'})
    
    if user['password_hash'] != hash_password(password):
        return jsonify({'success': False, 'error': '密码错误'})
    
    # 更新最后登录时间
    cursor.execute('UPDATE pm_user SET last_login = ? WHERE id = ?', 
                   (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user['id']))
    conn.commit()
    conn.close()
    
    # 创建会话
    token = generate_token()
    USER_SESSIONS[token] = {
        'user': {
            'id': user['id'],
            'username': user['username'],
            'person_id': user['person_id'],
            'person_name': user['person_name'],
            'role': user['role']
        },
        'expire_time': datetime.now() + timedelta(hours=SESSION_EXPIRE_HOURS)
    }
    
    return jsonify({
        'success': True,
        'token': token,
        'user': USER_SESSIONS[token]['user']
    })

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    """用户退出登录"""
    token = request.headers.get('Authorization', '')
    if token.startswith('Bearer '):
        token = token[7:]
    
    if token in USER_SESSIONS:
        del USER_SESSIONS[token]
    
    return jsonify({'success': True, 'message': '已退出登录'})

@app.route('/api/auth/me', methods=['GET'])
def get_current_user():
    """获取当前登录用户信息"""
    token = request.headers.get('Authorization', '')
    if token.startswith('Bearer '):
        token = token[7:]
    
    if not token or token not in USER_SESSIONS:
        return jsonify({'success': False, 'error': '未登录'})
    
    return jsonify({'success': True, 'user': USER_SESSIONS[token]['user']})

@app.route('/api/users', methods=['GET'])
@check_auth
def get_users():
    """获取所有用户列表（需要登录）"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT u.id, u.username, u.role, u.last_login, u.created_at, p.name as person_name
        FROM pm_user u
        LEFT JOIN person p ON u.person_id = p.id
        ORDER BY u.username
    ''')
    users = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify({'success': True, 'users': users})

@app.route('/api/users/change-password', methods=['POST'])
@check_auth
def change_password():
    """修改密码"""
    data = request.json
    old_password = data.get('old_password', '')
    new_password = data.get('new_password', '')
    
    # 密码强度验证
    if len(new_password) < 8:
        return jsonify({'success': False, 'error': '新密码长度至少8位'})
    if not re.search(r'[a-zA-Z]', new_password):
        return jsonify({'success': False, 'error': '新密码必须包含字母'})
    if not re.search(r'[0-9]', new_password):
        return jsonify({'success': False, 'error': '新密码必须包含数字'})
    if new_password == old_password:
        return jsonify({'success': False, 'error': '新密码不能与旧密码相同'})
    
    user = request.current_user
    conn = get_db()
    cursor = conn.cursor()
    
    # 验证旧密码
    cursor.execute('SELECT password_hash FROM pm_user WHERE id = ?', (user['id']))
    row = cursor.fetchone()
    
    if row['password_hash'] != hash_password(old_password):
        return jsonify({'success': False, 'error': '原密码错误'})
    
    # 更新密码
    cursor.execute('UPDATE pm_user SET password_hash = ? WHERE id = ?', 
                   (hash_password(new_password), user['id']))
    conn.commit()
    conn.close()
    
    # 记录操作日志
    add_log('user', user['id'], 'change_password', None, None, user.get('person_id'), user.get('person_name'))
    
    return jsonify({'success': True, 'message': '密码修改成功'})

@app.route('/api/users/init', methods=['POST'])
def init_users():
    """初始化用户账号（首次创建）"""
    count = create_users_from_persons()
    return jsonify({'success': True, 'message': f'成功创建 {count} 个用户账号', 'default_password': 'pm2026'})

# ========== 管理员 API ==========

@app.route('/api/admin/users/<int:user_id>/role', methods=['PUT'])
@check_auth
@require_admin
def update_user_role(user_id):
    """管理员：修改用户角色"""
    data = request.json
    new_role = data.get('role', 'user')
    
    if new_role not in ['admin', 'user']:
        return jsonify({'success': False, 'error': '无效的角色'})
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE pm_user SET role = ? WHERE id = ?', (new_role, user_id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': f'角色已更新为 {new_role}'})

@app.route('/api/admin/users/<int:user_id>/reset-password', methods=['POST'])
@check_auth
@require_admin
def admin_reset_password(user_id):
    """管理员：重置用户密码"""
    new_password = "pm2026"  # 重置为默认密码
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE pm_user SET password_hash = ? WHERE id = ?', 
                   (hash_password(new_password), user_id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': f'密码已重置为默认密码 {new_password}'})

@app.route('/api/admin/stats', methods=['GET'])
@check_auth
@require_admin
def admin_stats():
    """管理员：系统统计"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 项目统计
    cursor.execute('SELECT COUNT(*) FROM project')
    project_count = cursor.fetchone()[0]
    
    # 任务统计
    cursor.execute('SELECT COUNT(*) FROM task')
    task_count = cursor.fetchone()[0]
    
    # 用户统计
    cursor.execute('SELECT COUNT(*) FROM pm_user')
    user_count = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM pm_user WHERE role = "admin"')
    admin_count = cursor.fetchone()[0]
    
    # 问题统计
    cursor.execute('SELECT COUNT(*) FROM issue WHERE status != "resolved"')
    issue_count = cursor.fetchone()[0]
    
    # 最近登录
    cursor.execute('''
        SELECT u.username, p.name, u.last_login
        FROM pm_user u
        LEFT JOIN person p ON u.person_id = p.id
        WHERE u.last_login IS NOT NULL
        ORDER BY u.last_login DESC
        LIMIT 10
    ''')
    recent_logins = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    return jsonify({
        'success': True,
        'stats': {
            'projects': project_count,
            'tasks': task_count,
            'users': user_count,
            'admins': admin_count,
            'pending_issues': issue_count
        },
        'recent_logins': recent_logins
    })

# ========== 项目管理 API ==========

@app.route('/api/projects', methods=['GET'])
@check_auth
def get_projects():
    """获取项目列表（管理员看全部，普通用户看参与+被授权查看的项目）"""
    status = request.args.get('status', '')
    include_archived = request.args.get('include_archived', 'false')
    conn = get_db()
    cursor = conn.cursor()
    
    # 数据权限隔离
    current_user = getattr(request, 'current_user', {}) or {}
    is_admin = current_user.get('role') == 'admin'
    person_id = current_user.get('person_id')
    
    my_project_ids = get_user_project_ids(person_id, is_admin)
    
    if my_project_ids is not None:
        # 普通用户：只看参与+被授权查看的项目
        if not my_project_ids:
            conn.close()
            return jsonify({'success': True, 'projects': []})
        placeholders = ','.join(['?'] * len(my_project_ids))
        if status:
            cursor.execute(f'''
                SELECT p.*, per.name as owner_name
                FROM project p LEFT JOIN person per ON p.owner_id = per.id
                WHERE p.id IN ({placeholders}) AND p.status = ?
                ORDER BY p.updated_at DESC
            ''', my_project_ids + [status])
        elif include_archived == 'true':
            cursor.execute(f'''
                SELECT p.*, per.name as owner_name
                FROM project p LEFT JOIN person per ON p.owner_id = per.id
                WHERE p.id IN ({placeholders})
                ORDER BY p.updated_at DESC
            ''', my_project_ids)
        else:
            cursor.execute(f'''
                SELECT p.*, per.name as owner_name
                FROM project p LEFT JOIN person per ON p.owner_id = per.id
                WHERE p.id IN ({placeholders}) AND p.status != 'archived'
                ORDER BY p.updated_at DESC
            ''', my_project_ids)
    else:
        # 管理员：看全部
        if status:
            cursor.execute('''
                SELECT p.*, per.name as owner_name
                FROM project p LEFT JOIN person per ON p.owner_id = per.id
                WHERE p.status = ? ORDER BY p.updated_at DESC
            ''', (status,))
        elif include_archived == 'true':
            cursor.execute('''
                SELECT p.*, per.name as owner_name
                FROM project p LEFT JOIN person per ON p.owner_id = per.id
                ORDER BY p.updated_at DESC
            ''')
        else:
            cursor.execute('''
                SELECT p.*, per.name as owner_name
                FROM project p
                LEFT JOIN person per ON p.owner_id = per.id
                WHERE p.status != 'archived' ORDER BY p.updated_at DESC
            ''')
    
    projects = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    for p in projects:
        p['progress'] = calculate_project_progress(p['id'])
    
    return jsonify({'success': True, 'projects': projects})

@app.route('/api/projects', methods=['POST'])
@check_auth
def create_project():
    """创建项目（需要登录）"""
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO project (name, description, status, priority, start_date, target_end_date, owner_id)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (data.get('name'), data.get('description', ''), data.get('status', 'pending'), 
          data.get('priority', 'medium'), data.get('start_date'), 
          data.get('target_end_date'), data.get('owner_id')))
    
    project_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    # 记录操作日志
    add_log('project', project_id, 'create', None, data.get('name'), 
            request.current_user.get('person_id'), request.current_user.get('person_name'))
    
    return jsonify({'success': True, 'id': project_id, 'message': '项目创建成功'})

@app.route('/api/projects/<int:project_id>', methods=['GET'])
def get_project(project_id):
    """获取项目详情"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT p.*, per.name as owner_name
        FROM project p
        LEFT JOIN person per ON p.owner_id = per.id
        WHERE p.id = ?
    ''', (project_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'error': '项目不存在'}), 404
    
    project = dict(row)
    
    cursor.execute('''
        SELECT t.*, per.name as assignee_name
        FROM task t
        LEFT JOIN person per ON t.assignee_id = per.id
        WHERE t.project_id = ? ORDER BY t.due_date
    ''', (project_id,))
    tasks = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute('SELECT * FROM phase WHERE project_id = ? ORDER BY order_num', (project_id,))
    phases = [dict(row) for row in cursor.fetchall()]
    
    # 计算每个阶段的进度
    for phase in phases:
        phase['progress'] = calculate_phase_progress(phase['id'])
    
    cursor.execute('SELECT * FROM issue WHERE project_id = ? ORDER BY severity DESC, created_at DESC', (project_id,))
    issues = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    project['progress'] = calculate_project_progress(project_id)
    project['phases'] = phases
    project['tasks'] = tasks
    project['issues'] = issues
    
    return jsonify({'success': True, 'project': project})

@app.route('/api/projects/<int:project_id>', methods=['PUT'])
@check_auth
def update_project(project_id):
    """更新项目（需要登录）"""
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        UPDATE project SET 
            name = COALESCE(?, name),
            description = COALESCE(?, description),
            status = COALESCE(?, status),
            priority = COALESCE(?, priority),
            target_end_date = COALESCE(?, target_end_date),
            actual_end_date = COALESCE(?, actual_end_date),
            owner_id = COALESCE(?, owner_id),
            updated_at = ?
        WHERE id = ?
    ''', (data.get('name'), data.get('description'), data.get('status'),
          data.get('priority'), data.get('target_end_date'), data.get('actual_end_date'),
          data.get('owner_id'), datetime.now().isoformat(), project_id))
    
    conn.commit()
    conn.close()
    
    # 记录操作日志
    add_log('project', project_id, 'update', None, json.dumps(data, ensure_ascii=False), request.current_user.get('person_id'), request.current_user.get('person_name'))
    
    return jsonify({'success': True, 'message': '项目更新成功'})

@app.route('/api/projects/<int:project_id>', methods=['DELETE'])
@check_auth
@require_admin
def delete_project(project_id):
    """删除项目（需要管理员权限）"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 获取项目名称用于日志
    cursor.execute('SELECT name FROM project WHERE id = ?', (project_id,))
    project = cursor.fetchone()
    project_name = project['name'] if project else ''
    
    cursor.execute('DELETE FROM issue WHERE project_id = ?', (project_id,))
    cursor.execute('DELETE FROM reminder WHERE task_id IN (SELECT id FROM task WHERE project_id = ?)', (project_id,))
    cursor.execute('DELETE FROM task WHERE project_id = ?', (project_id,))
    cursor.execute('DELETE FROM phase WHERE project_id = ?', (project_id,))
    cursor.execute('DELETE FROM project WHERE id = ?', (project_id,))
    
    conn.commit()
    conn.close()
    
    # 记录操作日志
    add_log('project', project_id, 'delete', project_name, None, request.current_user.get('person_id'), request.current_user.get('person_name'))
    
    return jsonify({'success': True, 'message': '项目已删除'})

# ========== 项目结项 API ==========

@app.route('/api/projects/<int:project_id>/archive', methods=['POST'])
@check_auth
@require_admin
def archive_project(project_id):
    """结项项目（需要管理员权限）"""
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    # 获取项目信息
    cursor.execute('SELECT * FROM project WHERE id = ?', (project_id,))
    project = cursor.fetchone()
    if not project:
        conn.close()
        return jsonify({'success': False, 'error': '项目不存在'}), 404
    
    # 检查项目状态
    if project['status'] == 'archived':
        conn.close()
        return jsonify({'success': False, 'error': '项目已结项'}), 400
    
    # 更新项目状态为已结项
    archive_notes = data.get('archive_notes', '')
    cursor.execute('''
        UPDATE project SET 
            status = 'archived',
            archived_at = ?,
            archive_notes = ?,
            actual_end_date = COALESCE(?, actual_end_date),
            updated_at = ?
        WHERE id = ?
    ''', (datetime.now().isoformat(), archive_notes, 
          data.get('actual_end_date'), datetime.now().isoformat(), project_id))
    
    # 获取结项统计信息
    cursor.execute('SELECT COUNT(*) FROM task WHERE project_id = ? AND status = "completed"', (project_id,))
    completed_tasks = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM task WHERE project_id = ?', (project_id,))
    total_tasks = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM issue WHERE project_id = ? AND status = "resolved"', (project_id,))
    resolved_issues = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM issue WHERE project_id = ?', (project_id,))
    total_issues = cursor.fetchone()[0]
    
    conn.commit()
    conn.close()
    
    # 记录操作日志
    add_log('project', project_id, 'archive', project['name'], archive_notes, request.current_user.get('person_id'), request.current_user.get('person_name'))
    
    return jsonify({
        'success': True,
        'message': '项目已结项',
        'stats': {
            'completed_tasks': completed_tasks,
            'total_tasks': total_tasks,
            'resolved_issues': resolved_issues,
            'total_issues': total_issues
        }
    })

@app.route('/api/projects/<int:project_id>/unarchive', methods=['POST'])
@check_auth
@require_admin
def unarchive_project(project_id):
    """撤销结项（需要管理员权限）"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT status FROM project WHERE id = ?', (project_id,))
    project = cursor.fetchone()
    if not project:
        conn.close()
        return jsonify({'success': False, 'error': '项目不存在'}), 404
    
    if project['status'] != 'archived':
        conn.close()
        return jsonify({'success': False, 'error': '项目未结项'}), 400
    
    cursor.execute('''
        UPDATE project SET 
            status = 'completed',
            archived_at = NULL,
            updated_at = ?
        WHERE id = ?
    ''', (datetime.now().isoformat(), project_id))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': '已撤销结项'})

@app.route('/api/projects/archived', methods=['GET'])
def get_archived_projects():
    """获取已结项项目列表"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT p.*, per.name as owner_name
        FROM project p
        LEFT JOIN person per ON p.owner_id = per.id
        WHERE p.status = 'archived'
        ORDER BY p.archived_at DESC
    ''')
    
    projects = [dict(row) for row in cursor.fetchall()]
    
    # 获取每个结项项目的统计信息
    for p in projects:
        cursor.execute('SELECT COUNT(*) FROM task WHERE project_id = ? AND status = "completed"', (p['id'],))
        p['completed_tasks'] = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM task WHERE project_id = ?', (p['id'],))
        p['total_tasks'] = cursor.fetchone()[0]
    
    conn.close()
    
    return jsonify({'success': True, 'projects': projects})

# ========== 阶段管理 API ==========

@app.route('/api/phases', methods=['GET'])
@app.route('/api/phases', methods=['GET'])
@check_auth
def get_phases():
    """获取阶段列表（管理员看全部，普通用户看参与+授权查看项目的阶段）"""
    project_id = request.args.get('project_id')
    include_archived = request.args.get('include_archived', 'false')
    conn = get_db()
    cursor = conn.cursor()
    
    # 数据权限隔离
    current_user = getattr(request, 'current_user', {}) or {}
    is_admin = current_user.get('role') == 'admin'
    person_id = current_user.get('person_id')
    my_project_ids = get_user_project_ids(person_id, is_admin)
    
    if my_project_ids is not None:
        # 普通用户
        if not my_project_ids:
            conn.close()
            return jsonify({'success': True, 'phases': []})
        if project_id and int(project_id) not in my_project_ids:
            conn.close()
            return jsonify({'success': True, 'phases': []})
        
        placeholders = ','.join(['?'] * len(my_project_ids))
        if include_archived == 'true':
            cursor.execute(f'''
                SELECT * FROM phase WHERE project_id IN ({placeholders})
                ORDER BY project_id, order_num
            ''', my_project_ids)
        else:
            cursor.execute(f'''
                SELECT ph.* FROM phase ph
                LEFT JOIN project p ON ph.project_id = p.id
                WHERE ph.project_id IN ({placeholders}) AND p.status != "archived"
                ORDER BY ph.project_id, ph.order_num
            ''', my_project_ids)
    else:
        # 管理员看全部
        if project_id:
            cursor.execute('''
                SELECT ph.* FROM phase ph
                LEFT JOIN project p ON ph.project_id = p.id
                WHERE ph.project_id = ? AND (p.status != "archived" OR ? = "true")
                ORDER BY ph.order_num
            ''', (project_id, include_archived))
        else:
            if include_archived == 'true':
                cursor.execute('SELECT * FROM phase ORDER BY project_id, order_num')
            else:
                cursor.execute('''
                    SELECT ph.* FROM phase ph
                    LEFT JOIN project p ON ph.project_id = p.id
                    WHERE p.status != "archived"
                    ORDER BY ph.project_id, ph.order_num
                ''')
    
    phases = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return jsonify({'success': True, 'phases': phases})

@app.route('/api/phases', methods=['POST'])
@check_auth
def create_phase():
    """创建阶段（管理员或项目owner）"""
    data = request.json
    user = request.current_user
    person_id = user.get('person_id')
    
    if user.get('role') != 'admin' and not is_project_owner(data['project_id'], person_id):
        return jsonify({'success': False, 'error': '只有管理员或项目负责人可以创建阶段'}), 403
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO phase (project_id, name, order_num, start_date, end_date)
        VALUES (?, ?, ?, ?, ?)
    ''', (data['project_id'], data['name'], data.get('order_num', 0),
          data.get('start_date'), data.get('end_date')))
    
    phase_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'id': phase_id})

@app.route('/api/phases/<int:phase_id>', methods=['PUT'])
@check_auth
def update_phase(phase_id):
    """更新阶段（管理员或项目owner）"""
    data = request.json
    user = request.current_user
    person_id = user.get('person_id')
    
    # 获取阶段所属项目ID
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT project_id FROM phase WHERE id = ?', (phase_id,))
    phase = cursor.fetchone()
    if not phase:
        conn.close()
        return jsonify({'success': False, 'error': '阶段不存在'}), 404
    
    if user.get('role') != 'admin' and not is_project_owner(phase['project_id'], person_id):
        conn.close()
        return jsonify({'success': False, 'error': '只有管理员或项目负责人可以修改阶段'}), 403
    
    cursor.execute('''
        UPDATE phase SET 
            name = COALESCE(?, name),
            order_num = COALESCE(?, order_num),
            status = COALESCE(?, status),
            start_date = COALESCE(?, start_date),
            end_date = COALESCE(?, end_date),
            description = COALESCE(?, description),
            updated_at = ?
        WHERE id = ?
    ''', (data.get('name'), data.get('order_num'), data.get('status'),
          data.get('start_date'), data.get('end_date'), data.get('description'),
          datetime.now().isoformat(), phase_id))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

@app.route('/api/phases/<int:phase_id>', methods=['DELETE'])
@check_auth
def delete_phase(phase_id):
    """删除阶段（管理员或项目owner）"""
    user = request.current_user
    person_id = user.get('person_id')
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT project_id FROM phase WHERE id = ?', (phase_id,))
    phase = cursor.fetchone()
    if not phase:
        conn.close()
        return jsonify({'success': False, 'error': '阶段不存在'}), 404
    
    if user.get('role') != 'admin' and not is_project_owner(phase['project_id'], person_id):
        conn.close()
        return jsonify({'success': False, 'error': '只有管理员或项目负责人可以删除阶段'}), 403
    
    # 删除阶段下的任务
    cursor.execute('DELETE FROM task WHERE phase_id = ?', (phase_id,))
    # 删除阶段
    cursor.execute('DELETE FROM phase WHERE id = ?', (phase_id,))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'deleted': True})

@app.route('/api/phases/<int:phase_id>', methods=['GET'])
def get_phase(phase_id):
    """获取单个阶段详情"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM phase WHERE id = ?', (phase_id,))
    row = cursor.fetchone()
    
    conn.close()
    
    if row:
        return jsonify({'success': True, 'phase': dict(row)})
    return jsonify({'success': False, 'error': 'Phase not found'}), 404

# ========== 任务管理 API ==========

@app.route('/api/tasks', methods=['GET'])
@check_auth
def get_tasks():
    """获取任务列表（管理员看全部，普通用户看参与+授权查看项目的所有任务）"""
    project_id = request.args.get('project_id')
    phase_id = request.args.get('phase_id')
    status = request.args.get('status')
    include_archived = request.args.get('include_archived', 'false')
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = '''SELECT t.*, per.name as assignee_name, p.name as project_name, p.status as project_status
               FROM task t 
               LEFT JOIN person per ON t.assignee_id = per.id
               LEFT JOIN project p ON t.project_id = p.id
               WHERE 1=1'''
    params = []
    
    # 数据权限隔离
    current_user = getattr(request, 'current_user', {}) or {}
    is_admin = current_user.get('role') == 'admin'
    person_id = current_user.get('person_id')
    my_project_ids = get_user_project_ids(person_id, is_admin)
    
    if my_project_ids is not None:
        if not my_project_ids:
            conn.close()
            return jsonify({'success': True, 'tasks': []})
        placeholders = ','.join(['?'] * len(my_project_ids))
        query += f' AND t.project_id IN ({placeholders})'
        params.extend(my_project_ids)
    
    # 默认排除已结项项目的任务
    if include_archived != 'true':
        query += ' AND p.status != "archived"'
    
    if project_id:
        query += ' AND t.project_id = ?'
        params.append(project_id)
    if phase_id:
        query += ' AND t.phase_id = ?'
        params.append(phase_id)
    if status:
        query += ' AND t.status = ?'
        params.append(status)
    
    query += ' ORDER BY t.due_date'
    cursor.execute(query, params)
    
    tasks = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return jsonify({'success': True, 'tasks': tasks})

@app.route('/api/tasks', methods=['POST'])
@check_auth
def create_task():
    """创建任务（管理员或项目owner可创建）"""
    data = request.json
    user = request.current_user
    person_id = user.get('person_id')
    
    # 权限检查：管理员或项目owner
    if user.get('role') != 'admin' and not is_project_owner(data['project_id'], person_id):
        return jsonify({'success': False, 'error': '只有管理员或项目负责人可以创建任务'}), 403
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO task (project_id, phase_id, name, description, priority, due_date, assignee_id)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (data['project_id'], data.get('phase_id'), data['name'],
          data.get('description', ''), data.get('priority', 'medium'), 
          data.get('due_date'), data.get('assignee_id')))
    
    task_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    # 更新进度链
    update_progress_chain(data['project_id'], data.get('phase_id'))
    
    # 记录操作日志
    operator = getattr(request, 'current_user', {}) or {}
    add_log('task', task_id, 'create', None, data.get('name'), operator.get('person_id'), operator.get('person_name'))
    
    # 通知被分配人
    if data.get('assignee_id'):
        create_notification(
            data['assignee_id'], 'task_assign',
            f'新任务分配: {data.get("name")}',
            f'{operator.get("person_name", "系统")} 将任务「{data.get("name")}」分配给了你',
            'task', task_id
        )
    
    return jsonify({'success': True, 'id': task_id})

@app.route('/api/tasks/<int:task_id>', methods=['GET'])
def get_task(task_id):
    """获取任务详情"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT t.*, per.name as assignee_name, p.name as project_name
        FROM task t
        LEFT JOIN person per ON t.assignee_id = per.id
        LEFT JOIN project p ON t.project_id = p.id
        WHERE t.id = ?
    ''', (task_id,))
    row = cursor.fetchone()
    
    if not row:
        conn.close()
        return jsonify({'success': False, 'error': '任务不存在'}), 404
    
    task = dict(row)
    conn.close()
    
    return jsonify({'success': True, 'task': task})

@app.route('/api/tasks/<int:task_id>', methods=['PUT'])
@check_auth
def update_task(task_id):
    """更新任务（管理员、项目owner、或任务assignee可修改）"""
    data = request.json
    user = request.current_user
    person_id = user.get('person_id')
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 获取任务信息（用于进度联动）
    cursor.execute('SELECT project_id, phase_id, assignee_id FROM task WHERE id = ?', (task_id,))
    task_info = cursor.fetchone()
    if not task_info:
        conn.close()
        return jsonify({'success': False, 'error': '任务不存在'}), 404
    
    # 权限检查：管理员 or 项目owner or 任务assignee
    project_id = task_info['project_id']
    phase_id = task_info['phase_id']
    old_assignee_id = task_info['assignee_id']
    if (user.get('role') != 'admin' and 
        not is_project_owner(project_id, person_id) and 
        old_assignee_id != person_id):
        conn.close()
        return jsonify({'success': False, 'error': '只能修改自己负责的任务'}), 403
    
    # 如果进度设置为100，自动标记为已完成
    if data.get('progress') == 100:
        data['status'] = 'completed'
        data['completed_date'] = datetime.now().strftime('%Y-%m-%d')
    
    cursor.execute('''
        UPDATE task SET 
            name = COALESCE(?, name),
            description = COALESCE(?, description),
            status = COALESCE(?, status),
            priority = COALESCE(?, priority),
            due_date = COALESCE(?, due_date),
            progress = COALESCE(?, progress),
            completed_date = COALESCE(?, completed_date),
            notes = COALESCE(?, notes),
            assignee_id = COALESCE(?, assignee_id),
            phase_id = COALESCE(?, phase_id),
            updated_at = ?
        WHERE id = ?
    ''', (data.get('name'), data.get('description'), data.get('status'),
          data.get('priority'), data.get('due_date'), data.get('progress'),
          data.get('completed_date'), data.get('notes'), data.get('assignee_id'),
          data.get('phase_id'),
          datetime.now().isoformat(), task_id))
    
    conn.commit()
    conn.close()
    
    # 更新进度链（任务变更 → 阶段进度 → 项目进度）
    new_phase_id = data.get('phase_id', phase_id)
    project_progress = update_progress_chain(project_id, new_phase_id)
    
    # 记录操作日志
    operator = getattr(request, 'current_user', {}) or {}
    log_action = 'update'
    if data.get('progress') == 100:
        log_action = 'complete'
    elif data.get('status') == 'completed':
        log_action = 'complete'
    add_log('task', task_id, log_action, None, json.dumps(data, ensure_ascii=False), operator.get('person_id'), operator.get('person_name'))
    
    # 通知：分配变更
    new_assignee = data.get('assignee_id')
    if new_assignee and str(new_assignee) != str(old_assignee_id or ''):
        create_notification(
            new_assignee, 'task_assign',
            f'任务重新分配: {data.get("name", "")}',
            f'{operator.get("person_name", "系统")} 将任务分配给了你',
            'task', task_id
        )
    
    # 通知：任务完成 → 通知项目owner
    if log_action == 'complete':
        conn2 = get_db()
        cur2 = conn2.cursor()
        cur2.execute('SELECT owner_id FROM project WHERE id = ?', (project_id,))
        proj_row = cur2.fetchone()
        conn2.close()
        if proj_row and proj_row['owner_id'] and proj_row['owner_id'] != operator.get('person_id'):
            create_notification(
                proj_row['owner_id'], 'task_complete',
                f'任务完成: {data.get("name", "")}',
                f'{operator.get("person_name", "未知")} 完成了任务「{data.get("name", "")}」',
                'task', task_id
            )
    
    return jsonify({'success': True, 'project_progress': project_progress})

@app.route('/api/tasks/<int:task_id>', methods=['DELETE'])
@check_auth
def delete_task(task_id):
    """删除任务（管理员或项目owner可删除）"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 获取任务信息（用于进度联动）
    cursor.execute('SELECT project_id, phase_id, name FROM task WHERE id = ?', (task_id,))
    task_info = cursor.fetchone()
    if not task_info:
        conn.close()
        return jsonify({'success': False, 'error': '任务不存在'}), 404
    
    # 权限检查
    user = request.current_user
    person_id = user.get('person_id')
    if user.get('role') != 'admin' and not is_project_owner(task_info['project_id'], person_id):
        conn.close()
        return jsonify({'success': False, 'error': '只有管理员或项目负责人可以删除任务'}), 403
    
    task_name = task_info['name']
    
    cursor.execute('DELETE FROM reminder WHERE task_id = ?', (task_id,))
    cursor.execute('DELETE FROM task WHERE id = ?', (task_id,))
    
    conn.commit()
    conn.close()
    
    # 更新进度链
    if task_info:
        update_progress_chain(task_info['project_id'], task_info['phase_id'])
    
    # 记录操作日志
    add_log('task', task_id, 'delete', task_name, None, request.current_user.get('person_id'), request.current_user.get('person_name'))
    
    return jsonify({'success': True})

# ========== 问题管理 API ==========

@app.route('/api/issues', methods=['GET'])
@check_auth
def get_issues():
    """获取问题列表（管理员看全部，普通用户看参与+授权查看项目的问题）"""
    project_id = request.args.get('project_id')
    status = request.args.get('status')
    include_archived = request.args.get('include_archived', 'false')
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = '''SELECT i.*, per.name as assignee_name, p.name as project_name, p.status as project_status,
               ph.name as phase_name, ph.order_num as phase_order
               FROM issue i
               LEFT JOIN person per ON i.assignee_id = per.id
               LEFT JOIN project p ON i.project_id = p.id
               LEFT JOIN phase ph ON i.phase_id = ph.id
               WHERE 1=1'''
    params = []
    
    # 数据权限隔离
    current_user = getattr(request, 'current_user', {}) or {}
    is_admin = current_user.get('role') == 'admin'
    person_id = current_user.get('person_id')
    my_project_ids = get_user_project_ids(person_id, is_admin)
    
    if my_project_ids is not None:
        if not my_project_ids:
            conn.close()
            return jsonify({'success': True, 'issues': []})
        placeholders = ','.join(['?'] * len(my_project_ids))
        query += f' AND i.project_id IN ({placeholders})'
        params.extend(my_project_ids)
    
    # 默认排除已结项项目的问题
    if include_archived != 'true':
        query += ' AND p.status != "archived"'
    
    if project_id:
        query += ' AND i.project_id = ?'
        params.append(project_id)
    if status:
        query += ' AND i.status = ?'
        params.append(status)
    
    query += ' ORDER BY i.severity DESC, i.created_at DESC'
    cursor.execute(query, params)
    
    issues = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return jsonify({'success': True, 'issues': issues})

@app.route('/api/issues', methods=['POST'])
@check_auth
def create_issue():
    """创建问题（管理员、项目owner、项目参与者可创建）"""
    data = request.json
    user = request.current_user
    person_id = user.get('person_id')
    
    # 权限检查：管理员 or 项目owner or 项目参与者（非查看者）
    if user.get('role') != 'admin':
        if not is_project_participant(data['project_id'], person_id):
            # 查看者不能创建问题
            if is_project_viewer(data['project_id'], person_id):
                return jsonify({'success': False, 'error': '查看者不能创建问题，如需操作请联系项目负责人'}), 403
            return jsonify({'success': False, 'error': '只有项目参与者可以创建问题'}), 403
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO issue (project_id, task_id, phase_id, title, description, severity, assignee_id)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (data['project_id'], data.get('task_id'), data.get('phase_id'), data['title'],
          data.get('description', ''), data.get('severity', 'medium'), data.get('assignee_id')))
    
    issue_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'id': issue_id})

@app.route('/api/issues/<int:issue_id>', methods=['PUT'])
@check_auth
def update_issue(issue_id):
    """更新问题（管理员、项目owner、或问题assignee可修改）"""
    data = request.json
    user = request.current_user
    person_id = user.get('person_id')
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT project_id, assignee_id FROM issue WHERE id = ?', (issue_id,))
    issue_info = cursor.fetchone()
    if not issue_info:
        conn.close()
        return jsonify({'success': False, 'error': '问题不存在'}), 404
    
    # 权限检查：管理员 or 项目owner or 问题assignee
    if (user.get('role') != 'admin' and 
        not is_project_owner(issue_info['project_id'], person_id) and
        issue_info['assignee_id'] != person_id):
        conn.close()
        return jsonify({'success': False, 'error': '只能修改自己负责的问题'}), 403
    
    if data.get('status') == 'resolved':
        data['resolved_at'] = datetime.now().isoformat()
    
    cursor.execute('''
        UPDATE issue SET 
            title = COALESCE(?, title),
            description = COALESCE(?, description),
            severity = COALESCE(?, severity),
            status = COALESCE(?, status),
            solution = COALESCE(?, solution),
            resolved_at = COALESCE(?, resolved_at),
            notes = COALESCE(?, notes),
            assignee_id = COALESCE(?, assignee_id),
            phase_id = COALESCE(?, phase_id)
        WHERE id = ?
    ''', (data.get('title'), data.get('description'), data.get('severity'),
          data.get('status'), data.get('solution'), data.get('resolved_at'),
          data.get('notes'), data.get('assignee_id'), data.get('phase_id'), issue_id))
    
    conn.commit()
    conn.close()
    
    # 重要问题写入Memos
    if data.get('severity') == 'high' and data.get('status') == 'resolved':
        write_issue_to_memos(issue_id, data)
    
    return jsonify({'success': True})

@app.route('/api/issues/<int:issue_id>', methods=['GET'])
def get_issue(issue_id):
    """获取单个问题详情"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT i.*, p.name as project_name, per.name as assignee_name
        FROM issue i
        LEFT JOIN project p ON i.project_id = p.id
        LEFT JOIN person per ON i.assignee_id = per.id
        WHERE i.id = ?
    ''', (issue_id,))
    
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        return jsonify({'success': False, 'error': '问题不存在'}), 404
    
    issue = dict(row)
    return jsonify({'success': True, 'issue': issue})

@app.route('/api/issues/<int:issue_id>', methods=['DELETE'])
@check_auth
def delete_issue(issue_id):
    """删除问题（管理员或项目owner）"""
    user = request.current_user
    person_id = user.get('person_id')
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT project_id FROM issue WHERE id = ?', (issue_id,))
    issue = cursor.fetchone()
    if not issue:
        conn.close()
        return jsonify({'success': False, 'error': '问题不存在'}), 404
    
    if user.get('role') != 'admin' and not is_project_owner(issue['project_id'], person_id):
        conn.close()
        return jsonify({'success': False, 'error': '只有管理员或项目负责人可以删除问题'}), 403
    
    cursor.execute('DELETE FROM issue WHERE id = ?', (issue_id,))
    conn.commit()
    deleted = cursor.rowcount > 0
    conn.close()
    
    if deleted:
        return jsonify({'success': True, 'message': '问题已删除'})
    else:
        return jsonify({'success': False, 'error': '删除失败'}), 500

# ========== 项目查看者授权 API ==========

@app.route('/api/projects/<int:project_id>/viewers', methods=['GET'])
@check_auth
def get_project_viewers(project_id):
    """获取项目的查看者列表（管理员或项目owner可查看）"""
    user = request.current_user
    person_id = user.get('person_id')
    
    if user.get('role') != 'admin' and not is_project_owner(project_id, person_id):
        return jsonify({'success': False, 'error': '只有管理员或项目负责人可以查看查看者列表'}), 403
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT pv.*, per.name as person_name, per.position as person_position,
               u.username, granter.name as granted_by_name
        FROM project_viewer pv
        LEFT JOIN person per ON pv.person_id = per.id
        LEFT JOIN pm_user u ON u.person_id = pv.person_id
        LEFT JOIN person granter ON pv.granted_by = granter.id
        WHERE pv.project_id = ?
        ORDER BY pv.granted_at DESC
    ''', (project_id,))
    
    viewers = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return jsonify({'success': True, 'viewers': viewers})

@app.route('/api/projects/<int:project_id>/viewers', methods=['POST'])
@check_auth
def add_project_viewer(project_id):
    """添加项目查看者（管理员或项目owner可授权）"""
    user = request.current_user
    person_id = user.get('person_id')
    
    if user.get('role') != 'admin' and not is_project_owner(project_id, person_id):
        return jsonify({'success': False, 'error': '只有管理员或项目负责人可以授权查看权限'}), 403
    
    data = request.json
    viewer_person_id = data.get('person_id')
    if not viewer_person_id:
        return jsonify({'success': False, 'error': '缺少person_id'}), 400
    
    # 不能给已经是参与者的人授权（他们已经有权限了）
    if is_project_participant(project_id, viewer_person_id):
        return jsonify({'success': False, 'error': '该人员已是项目参与者，无需额外授权'}), 400
    
    # 不能给项目owner授权
    if is_project_owner(project_id, viewer_person_id):
        return jsonify({'success': False, 'error': '该人员已是项目负责人'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO project_viewer (project_id, person_id, granted_by)
            VALUES (?, ?, ?)
        ''', (project_id, viewer_person_id, person_id))
        conn.commit()
        viewer_id = cursor.lastrowid
        conn.close()
        
        # 记录操作日志
        add_log('project', project_id, 'add_viewer', f'授权查看权限给人员ID:{viewer_person_id}', None, person_id, user.get('person_name'))
        
        return jsonify({'success': True, 'id': viewer_id})
    except Exception as e:
        conn.close()
        if 'UNIQUE constraint' in str(e):
            return jsonify({'success': False, 'error': '该人员已有查看权限'}), 400
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/projects/<int:project_id>/viewers/<int:viewer_person_id>', methods=['DELETE'])
@check_auth
def remove_project_viewer(project_id, viewer_person_id):
    """移除项目查看者（管理员或项目owner可撤销）"""
    user = request.current_user
    person_id = user.get('person_id')
    
    if user.get('role') != 'admin' and not is_project_owner(project_id, person_id):
        return jsonify({'success': False, 'error': '只有管理员或项目负责人可以撤销查看权限'}), 403
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM project_viewer WHERE project_id = ? AND person_id = ?', (project_id, viewer_person_id))
    conn.commit()
    deleted = cursor.rowcount > 0
    conn.close()
    
    if deleted:
        add_log('project', project_id, 'remove_viewer', f'撤销人员ID:{viewer_person_id}的查看权限', None, person_id, user.get('person_name'))
        return jsonify({'success': True, 'message': '已撤销查看权限'})
    else:
        return jsonify({'success': False, 'error': '未找到该查看者记录'}), 404

# ========== 统计分析 API ==========

@app.route('/api/stats', methods=['GET'])
def get_stats():
    """获取统计数据（排除已结项项目）"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 项目统计（排除已结项）
    cursor.execute('SELECT COUNT(*) FROM project WHERE status != "archived"')
    total_projects = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM project WHERE status = "in_progress"')
    active_projects = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM project WHERE status = "completed"')
    completed_projects = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM project WHERE status = "archived"')
    archived_projects = cursor.fetchone()[0]
    
    # 任务统计（仅统计未结项项目的任务）
    cursor.execute('SELECT COUNT(*) FROM task WHERE project_id NOT IN (SELECT id FROM project WHERE status = "archived")')
    total_tasks = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM task WHERE status = "completed" AND project_id NOT IN (SELECT id FROM project WHERE status = "archived")')
    completed_tasks = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM task WHERE (status = "delayed" OR (status != "completed" AND due_date < ?)) AND project_id NOT IN (SELECT id FROM project WHERE status = "archived")', 
                   (datetime.now().strftime('%Y-%m-%d'),))
    delayed_tasks = cursor.fetchone()[0]
    
    # 问题统计（仅统计未结项项目的问题）
    cursor.execute('SELECT COUNT(*) FROM issue WHERE project_id NOT IN (SELECT id FROM project WHERE status = "archived")')
    total_issues = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM issue WHERE status = "open" AND project_id NOT IN (SELECT id FROM project WHERE status = "archived")')
    open_issues = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM issue WHERE status = "in_progress" AND project_id NOT IN (SELECT id FROM project WHERE status = "archived")')
    in_progress_issues = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM issue WHERE status = "resolved" AND project_id NOT IN (SELECT id FROM project WHERE status = "archived")')
    resolved_issues = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM issue WHERE status != "resolved" AND severity = "high" AND project_id NOT IN (SELECT id FROM project WHERE status = "archived")')
    high_priority_issues = cursor.fetchone()[0]
    
    conn.close()
    
    return jsonify({
        'success': True,
        'stats': {
            'projects': {
                'total': total_projects,
                'active': active_projects,
                'completed': completed_projects,
                'archived': archived_projects
            },
            'tasks': {
                'total': total_tasks,
                'completed': completed_tasks,
                'delayed': delayed_tasks,
                'completion_rate': round(completed_tasks / total_tasks * 100, 1) if total_tasks > 0 else 0
            },
            'issues': {
                'total': total_issues,
                'open': open_issues,
                'in_progress': in_progress_issues,
                'resolved': resolved_issues,
                'high_priority_open': high_priority_issues,
                'resolution_rate': round(resolved_issues / total_issues * 100, 1) if total_issues > 0 else 0
            }
        }
    })

# ========== 辅助函数 ==========

def calculate_phase_progress(phase_id):
    """计算阶段进度（基于该阶段下所有任务的完成情况）"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT COUNT(*) FROM task WHERE phase_id = ?', (phase_id,))
    total_tasks = cursor.fetchone()[0]
    
    if total_tasks == 0:
        return 0
    
    cursor.execute('SELECT COUNT(*) FROM task WHERE phase_id = ? AND status = "completed"', (phase_id,))
    completed_tasks = cursor.fetchone()[0]
    
    conn.close()
    
    return round(completed_tasks / total_tasks * 100, 1)

def calculate_project_progress(project_id):
    """计算项目进度（基于所有任务的完成情况）"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT COUNT(*) FROM task WHERE project_id = ?', (project_id,))
    total_tasks = cursor.fetchone()[0]
    
    if total_tasks == 0:
        return 0
    
    cursor.execute('SELECT COUNT(*) FROM task WHERE project_id = ? AND status = "completed"', (project_id,))
    completed_tasks = cursor.fetchone()[0]
    
    conn.close()
    
    return round(completed_tasks / total_tasks * 100, 1)

def update_progress_chain(project_id, phase_id=None):
    """更新进度链：任务变更 → 更新阶段进度 → 更新项目进度"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 如果有phase_id，更新阶段进度
    if phase_id:
        phase_progress = calculate_phase_progress(phase_id)
        cursor.execute('UPDATE phase SET progress = ?, updated_at = ? WHERE id = ?',
                      (phase_progress, datetime.now().isoformat(), phase_id))
    
    # 更新项目进度
    project_progress = calculate_project_progress(project_id)
    cursor.execute('UPDATE project SET progress = ?, updated_at = ? WHERE id = ?',
                  (project_progress, datetime.now().isoformat(), project_id))
    
    conn.commit()
    conn.close()
    
    return project_progress

def write_issue_to_memos(issue_id, data):
    """将重要问题写入Memos"""
    content = f"#项目管理 #问题解决\n\n问题: {data.get('title')}\n解决方案: {data.get('solution')}"
    subprocess.run(['python3', MEMOS_SCRIPT, content], capture_output=True)

def send_dingtalk_notification(message):
    """发送钉钉通知"""
    try:
        # 使用copaw发送钉钉消息
        result = subprocess.run(
            ['copaw', 'notify', '--channel', 'dingtalk', '--message', message],
            capture_output=True, text=True, timeout=30
        )
        return result.returncode == 0
    except Exception as e:
        print(f"钉钉通知发送失败: {e}")
        return False

# ========== 提醒管理 API ==========

@app.route('/api/reminders', methods=['GET'])
def get_reminders():
    """获取提醒列表"""
    status = request.args.get('status', '')
    task_id = request.args.get('task_id', '')
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = 'SELECT r.*, t.name as task_name, t.due_date as task_due_date FROM reminder r LEFT JOIN task t ON r.task_id = t.id WHERE 1=1'
    params = []
    
    if status:
        query += ' AND r.status = ?'
        params.append(status)
    if task_id:
        query += ' AND r.task_id = ?'
        params.append(task_id)
    
    query += ' ORDER BY r.reminder_date ASC'
    cursor.execute(query, params)
    
    reminders = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return jsonify({'success': True, 'reminders': reminders})

@app.route('/api/reminders', methods=['POST'])
def create_reminder():
    """创建提醒"""
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO reminder (task_id, reminder_type, reminder_date, status)
        VALUES (?, ?, ?, 'pending')
    ''', (data.get('task_id'), data.get('reminder_type'), data.get('reminder_date')))
    
    reminder_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'id': reminder_id})

@app.route('/api/reminders/<int:reminder_id>', methods=['PUT'])
def update_reminder(reminder_id):
    """更新提醒状态"""
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        UPDATE reminder SET status = ?, sent_at = ? WHERE id = ?
    ''', (data.get('status', 'sent'), datetime.now().isoformat(), reminder_id))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

@app.route('/api/reminders/check', methods=['GET'])
def check_due_reminders():
    """检查需要发送的提醒"""
    today = datetime.now().strftime('%Y-%m-%d')
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 查找今日需要发送的提醒
    cursor.execute('''
        SELECT r.*, t.name as task_name, p.name as project_name
        FROM reminder r
        LEFT JOIN task t ON r.task_id = t.id
        LEFT JOIN project p ON t.project_id = p.id
        WHERE r.reminder_date = ? AND r.status = 'pending'
    ''', (today))
    
    reminders = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return jsonify({'success': True, 'reminders': reminders, 'count': len(reminders)})

@app.route('/api/reminders/send', methods=['POST'])
def send_reminders():
    """发送到期提醒"""
    today = datetime.now().strftime('%Y-%m-%d')
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 查找今日需要发送的提醒
    cursor.execute('''
        SELECT r.*, t.name as task_name, p.name as project_name
        FROM reminder r
        LEFT JOIN task t ON r.task_id = t.id
        LEFT JOIN project p ON t.project_id = p.id
        WHERE r.reminder_date = ? AND r.status = 'pending'
    ''', (today))
    
    reminders = [dict(row) for row in cursor.fetchall()]
    
    sent_count = 0
    for r in reminders:
        msg = f"📋 项目提醒\n\n任务: {r['task_name']}\n项目: {r['project_name']}\n提醒类型: {r['reminder_type']}\n\n请及时处理！"
        if send_dingtalk_notification(msg):
            cursor.execute('UPDATE reminder SET status = ?, sent_at = ? WHERE id = ?', 
                          ('sent', datetime.now().isoformat(), r['id']))
            sent_count += 1
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'sent': sent_count, 'total': len(reminders)})

@app.route('/api/tasks/due-check', methods=['GET'])
def check_due_tasks():
    """检查即将到期和已延期的任务（排除已结项项目）"""
    today = datetime.now().strftime('%Y-%m-%d')
    tomorrow = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
    next_week = (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 已延期任务（排除已结项项目）
    cursor.execute('''
        SELECT t.*, p.name as project_name
        FROM task t LEFT JOIN project p ON t.project_id = p.id
        WHERE t.due_date < ? AND t.status != 'completed' AND p.status != 'archived'
        ORDER BY t.due_date ASC
    ''', [today])
    overdue_tasks = [dict(row) for row in cursor.fetchall()]
    
    # 今日到期任务（排除已结项项目）
    cursor.execute('''
        SELECT t.*, p.name as project_name
        FROM task t LEFT JOIN project p ON t.project_id = p.id
        WHERE t.due_date = ? AND t.status != 'completed' AND p.status != 'archived'
    ''', [today])
    today_tasks = [dict(row) for row in cursor.fetchall()]
    
    # 一周内到期任务（排除已结项项目）
    cursor.execute('''
        SELECT t.*, p.name as project_name
        FROM task t LEFT JOIN project p ON t.project_id = p.id
        WHERE t.due_date BETWEEN ? AND ? AND t.status != 'completed' AND p.status != 'archived'
        ORDER BY t.due_date ASC
    ''', [tomorrow, next_week])
    upcoming_tasks = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    return jsonify({
        'success': True,
        'overdue': overdue_tasks,
        'today': today_tasks,
        'upcoming': upcoming_tasks,
        'stats': {
            'overdue_count': len(overdue_tasks),
            'today_count': len(today_tasks),
            'upcoming_count': len(upcoming_tasks)
        }
    })

@app.route('/api/notify/due-tasks', methods=['POST'])
def notify_due_tasks():
    """发送到期任务通知到钉钉"""
    result = check_due_tasks()
    data = result.get_json()
    
    overdue = data.get('overdue', [])
    today = data.get('today', [])
    
    if not overdue and not today:
        return jsonify({'success': True, 'message': '无需要通知的任务'})
    
    # 构建通知消息
    msg_lines = ["📋 任务到期提醒\n\n"]
    
    if overdue:
        msg_lines.append("⚠️ 已延期任务:\n")
        for t in overdue[:5]:
            msg_lines.append(f"  • {t['name']} ({t['project_name']}) - 截止: {t['due_date']}\n")
        msg_lines.append("\n")
    
    if today:
        msg_lines.append("📅 今日到期:\n")
        for t in today[:5]:
            msg_lines.append(f"  • {t['name']} ({t['project_name']})\n")
    
    msg = ''.join(msg_lines)
    
    if send_dingtalk_notification(msg):
        return jsonify({'success': True, 'sent': True, 'message': msg})
    
    return jsonify({'success': False, 'error': '发送失败'})

# ========== 前端页面 ==========

@app.route('/', methods=['GET'])
def index():
    """返回前端页面"""
    return send_file(os.path.join(FRONTEND_DIR, 'index.html'))

@app.route('/<path:path>', methods=['GET'])
def static_files(path):
    """返回静态文件"""
    if path.startswith('api/'):
        # API路由不处理
        return jsonify({'error': 'Not found'}), 404
    return send_file(os.path.join(FRONTEND_DIR, path))

# ========== 人员管理 API ==========

@app.route('/api/persons', methods=['GET'])
@check_auth
def get_persons():
    """获取人员列表（管理员看全部，普通用户看参与项目相关人员）"""
    status = request.args.get('status', '')
    
    current_user = getattr(request, 'current_user', {}) or {}
    is_admin = current_user.get('role') == 'admin'
    
    conn = get_db()
    cursor = conn.cursor()
    
    if is_admin:
        # 管理员看全部
        if status:
            cursor.execute('SELECT * FROM person WHERE status = ? ORDER BY name', (status,))
        else:
            cursor.execute('SELECT * FROM person ORDER BY name')
    else:
        # 普通用户：看参与项目相关人员（owner、assignee、viewer）+管理员
        person_id = current_user.get('person_id')
        my_project_ids = get_user_project_ids(person_id, False)
        
        if not my_project_ids:
            # 没有参与任何项目，只看管理员和自己
            cursor.execute('''
                SELECT DISTINCT p.* FROM person p
                LEFT JOIN pm_user u ON u.person_id = p.id
                WHERE u.role = 'admin' OR p.id = ?
                ORDER BY p.name
            ''', (person_id,))
        else:
            placeholders = ','.join(['?'] * len(my_project_ids))
            # 看参与项目的owner、assignee、viewer + 管理员
            cursor.execute(f'''
                SELECT DISTINCT p.* FROM person p
                LEFT JOIN pm_user u ON u.person_id = p.id
                LEFT JOIN task t ON t.project_id IN ({placeholders}) AND t.assignee_id = p.id
                LEFT JOIN project_viewer pv ON pv.project_id IN ({placeholders}) AND pv.person_id = p.id
                LEFT JOIN project proj ON proj.id IN ({placeholders}) AND proj.owner_id = p.id
                WHERE u.role = 'admin' OR p.id = ? OR t.id IS NOT NULL OR pv.id IS NOT NULL OR proj.id IS NOT NULL
                ORDER BY p.name
            ''', my_project_ids + my_project_ids + my_project_ids + [person_id])
    
    persons = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return jsonify({'success': True, 'persons': persons})

@app.route('/api/persons', methods=['POST'])
@check_auth
def create_person():
    """创建人员（仅管理员）"""
    user = request.current_user
    if user.get('role') != 'admin':
        return jsonify({'success': False, 'error': '只有管理员可以创建人员'}), 403
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO person (employee_id, name, role, email, phone, department, job_sequence, line, sub_line, position, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (data.get('employee_id'), data.get('name'), data.get('role', 'member'), 
          data.get('email'), data.get('phone'), 
          data.get('department', '智能与数字化中心'),
          data.get('job_sequence'), data.get('line'), data.get('sub_line'), 
          data.get('position'), data.get('status', 'active')))
    
    person_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'id': person_id})

@app.route('/api/persons/<int:person_id>', methods=['GET'])
def get_person(person_id):
    """获取人员详情"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM person WHERE id = ?', (person_id,))
    row = cursor.fetchone()
    
    if row:
        # 获取该人员负责的项目数和任务数
        cursor.execute('SELECT COUNT(*) FROM project WHERE owner_id = ?', (person_id,))
        project_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM task WHERE assignee_id = ?', (person_id,))
        task_count = cursor.fetchone()[0]
        
        person = dict(row)
        person['project_count'] = project_count
        person['task_count'] = task_count
        conn.close()
        return jsonify({'success': True, 'person': person})
    
    conn.close()
    return jsonify({'success': False, 'error': 'Person not found'}), 404

@app.route('/api/persons/<int:person_id>', methods=['PUT'])
@check_auth
def update_person(person_id):
    """更新人员（仅管理员）"""
    user = request.current_user
    if user.get('role') != 'admin':
        return jsonify({'success': False, 'error': '只有管理员可以修改人员信息'}), 403
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        UPDATE person SET 
            employee_id = COALESCE(?, employee_id),
            name = COALESCE(?, name),
            role = COALESCE(?, role),
            email = COALESCE(?, email),
            phone = COALESCE(?, phone),
            department = COALESCE(?, department),
            job_sequence = COALESCE(?, job_sequence),
            line = COALESCE(?, line),
            sub_line = COALESCE(?, sub_line),
            position = COALESCE(?, position),
            status = COALESCE(?, status),
            updated_at = ?
        WHERE id = ?
    ''', (data.get('employee_id'), data.get('name'), data.get('role'), 
          data.get('email'), data.get('phone'), data.get('department'),
          data.get('job_sequence'), data.get('line'), data.get('sub_line'),
          data.get('position'), data.get('status'),
          datetime.now().isoformat(), person_id))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

@app.route('/api/persons/<int:person_id>', methods=['DELETE'])
@check_auth
def delete_person(person_id):
    """删除人员（仅管理员）"""
    user = request.current_user
    if user.get('role') != 'admin':
        return jsonify({'success': False, 'error': '只有管理员可以删除人员'}), 403
    conn = get_db()
    cursor = conn.cursor()
    
    # 检查是否有关联的项目或任务
    cursor.execute('SELECT COUNT(*) FROM project WHERE owner_id = ?', (person_id,))
    project_count = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM task WHERE assignee_id = ?', (person_id,))
    task_count = cursor.fetchone()[0]
    
    if project_count > 0 or task_count > 0:
        conn.close()
        return jsonify({'success': False, 'error': f'该人员负责{project_count}个项目和{task_count}个任务，无法删除'}), 400
    
    cursor.execute('DELETE FROM person WHERE id = ?', (person_id,))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'deleted': True})

# ========== 操作日志 ==========

def init_log_table():
    """初始化操作日志表"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS operation_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entity_type TEXT NOT NULL,
            entity_id INTEGER NOT NULL,
            action TEXT NOT NULL,
            old_value TEXT,
            new_value TEXT,
            operator_id INTEGER,
            operator_name TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

def add_log(entity_type, entity_id, action, old_value=None, new_value=None, operator_id=None, operator_name=None):
    """添加操作日志"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO operation_log (entity_type, entity_id, action, old_value, new_value, operator_id, operator_name)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (entity_type, entity_id, action, old_value, new_value, operator_id, operator_name))
    conn.commit()
    conn.close()

@app.route('/api/logs', methods=['GET'])
@check_auth
def get_logs():
    """获取操作日志（需要登录）"""
    entity_type = request.args.get('entity_type', '')
    entity_id = request.args.get('entity_id', '')
    operator_name = request.args.get('operator_name', '')
    action = request.args.get('action', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    limit = request.args.get('limit', '100')
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = 'SELECT * FROM operation_log WHERE 1=1'
    params = []
    
    if entity_type:
        query += ' AND entity_type = ?'
        params.append(entity_type)
    if entity_id:
        query += ' AND entity_id = ?'
        params.append(int(entity_id))
    if operator_name:
        query += ' AND operator_name LIKE ?'
        params.append(f'%{operator_name}%')
    if action:
        query += ' AND action = ?'
        params.append(action)
    if start_date:
        query += ' AND created_at >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND created_at <= ?'
        params.append(end_date)
    
    query += ' ORDER BY created_at DESC LIMIT ?'
    params.append(int(limit))
    
    cursor.execute(query, params)
    logs = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return jsonify({'success': True, 'logs': logs})

# ========== 评论功能 ==========

def init_comment_table():
    """初始化评论表"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS comment (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entity_type TEXT NOT NULL,
            entity_id INTEGER NOT NULL,
            content TEXT NOT NULL,
            author_id INTEGER,
            author_name TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

@app.route('/api/comments', methods=['GET'])
def get_comments():
    """获取评论列表"""
    entity_type = request.args.get('entity_type', '')
    entity_id = request.args.get('entity_id', '')
    
    conn = get_db()
    cursor = conn.cursor()
    
    if entity_type and entity_id:
        cursor.execute('''
            SELECT * FROM comment 
            WHERE entity_type = ? AND entity_id = ?
            ORDER BY created_at DESC
        ''', (entity_type, int(entity_id)))
        comments = [dict(row) for row in cursor.fetchall()]
    else:
        comments = []
    
    conn.close()
    return jsonify({'success': True, 'comments': comments})

@app.route('/api/comments', methods=['POST'])
def create_comment():
    """创建评论"""
    data = request.json
    entity_type = data.get('entity_type')
    entity_id = data.get('entity_id')
    content = data.get('content')
    author_id = data.get('author_id')
    author_name = data.get('author_name', '匿名')
    
    if not entity_type or not entity_id or not content:
        return jsonify({'success': False, 'error': '缺少必要参数'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO comment (entity_type, entity_id, content, author_id, author_name)
        VALUES (?, ?, ?, ?, ?)
    ''', (entity_type, entity_id, content, author_id, author_name))
    comment_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    # 添加日志
    add_log(entity_type, entity_id, 'comment_added', None, content, author_id, author_name)
    
    # 通知：评论任务时通知任务负责人
    if entity_type == 'task' and entity_id:
        try:
            conn2 = get_db()
            cursor2 = conn2.cursor()
            cursor2.execute('SELECT assignee_id, name FROM task WHERE id = ?', (entity_id,))
            task_row = cursor2.fetchone()
            conn2.close()
            if task_row and task_row['assignee_id'] and task_row['assignee_id'] != author_id:
                create_notification(
                    task_row['assignee_id'], 'comment',
                    f'新评论: {task_row["name"]}',
                    f'{author_name} 在任务「{task_row["name"]}」中发表了评论',
                    'task', entity_id
                )
        except Exception:
            pass
    
    return jsonify({'success': True, 'comment': {'id': comment_id, 'content': content}})

@app.route('/api/comments/<int:comment_id>', methods=['DELETE'])
def delete_comment(comment_id):
    """删除评论"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM comment WHERE id = ?', (comment_id,))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'deleted': True})

# ========== 模板功能 ==========

def init_template_table():
    """初始化模板表"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS template (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            type TEXT NOT NULL,
            description TEXT,
            content TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

@app.route('/api/templates', methods=['GET'])
def get_templates():
    """获取模板列表"""
    template_type = request.args.get('type', '')
    
    conn = get_db()
    cursor = conn.cursor()
    
    if template_type:
        cursor.execute('SELECT * FROM template WHERE type = ? ORDER BY name', (template_type,))
    else:
        cursor.execute('SELECT * FROM template ORDER BY type, name')
    
    templates = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return jsonify({'success': True, 'templates': templates})

@app.route('/api/templates', methods=['POST'])
def create_template():
    """创建模板"""
    data = request.json
    name = data.get('name')
    template_type = data.get('type', 'project')
    description = data.get('description', '')
    content = data.get('content', '{}')
    
    if not name:
        return jsonify({'success': False, 'error': '模板名称不能为空'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO template (name, type, description, content)
        VALUES (?, ?, ?, ?)
    ''', (name, template_type, description, content))
    template_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'template': {'id': template_id, 'name': name}})

@app.route('/api/templates/<int:template_id>', methods=['GET'])
def get_template(template_id):
    """获取单个模板"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM template WHERE id = ?', (template_id,))
    row = cursor.fetchone()
    conn.close()
    
    if row:
        return jsonify({'success': True, 'template': dict(row)})
    return jsonify({'success': False, 'error': '模板不存在'}), 404

@app.route('/api/templates/<int:template_id>', methods=['PUT'])
def update_template(template_id):
    """更新模板"""
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    updates = []
    values = []
    for field in ['name', 'type', 'description', 'content']:
        if field in data:
            updates.append(f'{field} = ?')
            values.append(data[field])
    
    if updates:
        values.append(template_id)
        cursor.execute(f'UPDATE template SET {", ".join(updates)}, updated_at = CURRENT_TIMESTAMP WHERE id = ?', values)
        conn.commit()
    
    conn.close()
    return jsonify({'success': True})

@app.route('/api/templates/<int:template_id>', methods=['DELETE'])
def delete_template(template_id):
    """删除模板"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM template WHERE id = ?', (template_id,))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'deleted': True})

@app.route('/api/templates/apply', methods=['POST'])
def apply_template():
    """应用模板创建项目或任务"""
    data = request.json
    template_id = data.get('template_id')
    name = data.get('name')
    
    if not template_id or not name:
        return jsonify({'success': False, 'error': '缺少必要参数'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 获取模板内容
    cursor.execute('SELECT * FROM template WHERE id = ?', (template_id,))
    template = cursor.fetchone()
    
    if not template:
        conn.close()
        return jsonify({'success': False, 'error': '模板不存在'}), 404
    
    import json
    content = json.loads(template['content'])
    
    # 根据模板类型创建
    if template['type'] == 'project':
        # 创建项目
        cursor.execute('''
            INSERT INTO project (name, description, status, priority, owner_id, start_date, target_end_date)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            name,
            content.get('description', ''),
            'pending',
            content.get('priority', 'medium'),
            content.get('owner_id'),
            datetime.now().strftime('%Y-%m-%d'),
            content.get('target_end_date')
        ))
        project_id = cursor.lastrowid
        
        # 创建预设阶段
        for phase in content.get('phases', []):
            cursor.execute('''
                INSERT INTO phase (project_id, name, start_date, end_date, order_num, status)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (project_id, phase.get('name'), phase.get('start_date'), phase.get('end_date'), phase.get('order_num', 1), 'pending'))
        
        # 创建预设任务
        for task in content.get('tasks', []):
            cursor.execute('''
                INSERT INTO task (project_id, name, description, status, priority, due_date)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (project_id, task.get('name'), task.get('description', ''), 'pending', task.get('priority', 'medium'), task.get('due_date')))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'project_id': project_id})
    
    elif template['type'] == 'task':
        # 创建任务
        project_id = data.get('project_id')
        cursor.execute('''
            INSERT INTO task (project_id, name, description, status, priority, due_date)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            project_id,
            name,
            content.get('description', ''),
            'pending',
            content.get('priority', 'medium'),
            content.get('due_date')
        ))
        task_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'task_id': task_id})
    
    conn.close()
    return jsonify({'success': False, 'error': '未知模板类型'}), 400

# ========== 周报/月报生成 ==========

@app.route('/api/report/weekly', methods=['GET'])
def get_weekly_report():
    """生成周报"""
    from datetime import datetime, timedelta
    
    # 计算本周起止时间
    today = datetime.now()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 本周完成的任务
    cursor.execute('''
        SELECT t.*, p.name as project_name, per.name as assignee_name
        FROM task t
        LEFT JOIN project p ON t.project_id = p.id
        LEFT JOIN person per ON t.assignee_id = per.id
        WHERE t.completed_date >= ? AND t.completed_date <= ?
        ORDER BY t.completed_date DESC
    ''', (week_start.strftime('%Y-%m-%d'), week_end.strftime('%Y-%m-%d')))
    completed_tasks = [dict(row) for row in cursor.fetchall()]
    
    # 本周创建的任务
    cursor.execute('''
        SELECT t.*, p.name as project_name
        FROM task t
        LEFT JOIN project p ON t.project_id = p.id
        WHERE t.created_at >= ?
        ORDER BY t.created_at DESC
    ''', (week_start.strftime('%Y-%m-%d') + ' 00:00:00',))
    new_tasks = [dict(row) for row in cursor.fetchall()]
    
    # 本周解决的问题
    cursor.execute('''
        SELECT i.*, p.name as project_name
        FROM issue i
        LEFT JOIN project p ON i.project_id = p.id
        WHERE i.resolved_at >= ? AND i.status = 'resolved'
        ORDER BY i.resolved_at DESC
    ''', (week_start.strftime('%Y-%m-%d') + ' 00:00:00',))
    resolved_issues = [dict(row) for row in cursor.fetchall()]
    
    # 进行中的项目
    cursor.execute('''
        SELECT p.*, per.name as owner_name
        FROM project p
        LEFT JOIN person per ON p.owner_id = per.id
        WHERE p.status = 'in_progress'
        ORDER BY p.progress DESC
    ''')
    active_projects = [dict(row) for row in cursor.fetchall()]
    
    # 统计数据
    stats = {
        'week_start': week_start.strftime('%Y-%m-%d'),
        'week_end': week_end.strftime('%Y-%m-%d'),
        'completed_tasks': len(completed_tasks),
        'new_tasks': len(new_tasks),
        'resolved_issues': len(resolved_issues),
        'active_projects': len(active_projects),
        'total_tasks': cursor.execute('SELECT COUNT(*) FROM task').fetchone()[0],
        'completed_total': cursor.execute('SELECT COUNT(*) FROM task WHERE status="completed"').fetchone()[0],
    }
    
    conn.close()
    
    return jsonify({
        'success': True,
        'report': {
            'type': 'weekly',
            'period': f'{week_start.strftime("%Y-%m-%d")} ~ {week_end.strftime("%Y-%m-%d")}',
            'stats': stats,
            'completed_tasks': completed_tasks,
            'new_tasks': new_tasks,
            'resolved_issues': resolved_issues,
            'active_projects': active_projects
        }
    })

@app.route('/api/report/monthly', methods=['GET'])
def get_monthly_report():
    """生成月报"""
    from datetime import datetime
    
    today = datetime.now()
    month_start = today.replace(day=1)
    month_end = today.replace(day=28)  # 简化处理
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 本月完成的任务
    cursor.execute('''
        SELECT t.*, p.name as project_name, per.name as assignee_name
        FROM task t
        LEFT JOIN project p ON t.project_id = p.id
        LEFT JOIN person per ON t.assignee_id = per.id
        WHERE t.completed_date >= ? AND t.completed_date <= ?
        ORDER BY t.completed_date DESC
    ''', (month_start.strftime('%Y-%m-%d'), month_end.strftime('%Y-%m-%d')))
    completed_tasks = [dict(row) for row in cursor.fetchall()]
    
    # 本月完成的项目
    cursor.execute('''
        SELECT p.*, per.name as owner_name
        FROM project p
        LEFT JOIN person per ON p.owner_id = per.id
        WHERE p.actual_end_date >= ? AND p.actual_end_date <= ?
        ORDER BY p.actual_end_date DESC
    ''', (month_start.strftime('%Y-%m-%d'), month_end.strftime('%Y-%m-%d')))
    completed_projects = [dict(row) for row in cursor.fetchall()]
    
    # 本月解决的问题
    cursor.execute('''
        SELECT COUNT(*) FROM issue WHERE resolved_at >= ? AND status = 'resolved'
    ''', (month_start.strftime('%Y-%m-%d') + ' 00:00:00',))
    resolved_count = cursor.fetchone()[0]
    
    # 人员绩效
    cursor.execute('''
        SELECT per.name, COUNT(t.id) as task_count,
               SUM(CASE WHEN t.status = 'completed' THEN 1 ELSE 0 END) as completed_count
        FROM person per
        LEFT JOIN task t ON t.assignee_id = per.id
        GROUP BY per.id
        ORDER BY completed_count DESC
    ''')
    person_stats = [dict(row) for row in cursor.fetchall()]
    
    stats = {
        'month': today.strftime('%Y-%m'),
        'completed_tasks': len(completed_tasks),
        'completed_projects': len(completed_projects),
        'resolved_issues': resolved_count,
    }
    
    conn.close()
    
    return jsonify({
        'success': True,
        'report': {
            'type': 'monthly',
            'period': today.strftime('%Y年%m月'),
            'stats': stats,
            'completed_tasks': completed_tasks,
            'completed_projects': completed_projects,
            'person_stats': person_stats
        }
    })

@app.route('/api/report/export', methods=['POST'])
def export_report():
    """导出报告"""
    data = request.json
    report_type = data.get('type', 'weekly')
    format_type = data.get('format', 'html')
    
    # 获取报告数据
    if report_type == 'weekly':
        from datetime import datetime, timedelta
        today = datetime.now()
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT t.*, p.name as project_name FROM task t
            LEFT JOIN project p ON t.project_id = p.id
            WHERE t.status = 'completed'
            ORDER BY t.completed_date DESC LIMIT 20
        ''')
        tasks = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        
        content = f"""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><title>周报 - {week_start.strftime('%Y-%m-%d')}</title></head>
<body style="font-family: sans-serif; padding: 20px;">
<h1>📊 项目管理周报</h1>
<h2>报告周期：{week_start.strftime('%Y-%m-%d')} ~ {week_end.strftime('%Y-%m-%d')}</h2>
<hr>
<h3>✅ 已完成任务 ({len(tasks)}个)</h3>
<table border="1" style="border-collapse: collapse; width: 100%;">
<tr><th>任务名称</th><th>所属项目</th><th>完成日期</th></tr>
{"".join([f"<tr><td>{t['name']}</td><td>{t['project_name'] or '-'}</td><td>{t['completed_date'] or '-'}</td></tr>" for t in tasks])}
</table>
<hr>
<p style="color: #666;">生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
</body>
</html>
"""
        
        filename = f"weekly_report_{week_start.strftime('%Y%m%d')}.html"
        
    else:
        content = "<html><body><h1>月报</h1></body></html>"
        filename = "monthly_report.html"
    
    import tempfile
    import os
    filepath = os.path.join(tempfile.gettempdir(), filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(content)
    
    return send_file(filepath, as_attachment=True, download_name=filename)

# ========== 附件管理 ==========

UPLOAD_DIR = os.environ.get('PM_UPLOAD_DIR', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'attachments'))
os.makedirs(UPLOAD_DIR, exist_ok=True)

def init_attachment_table():
    """初始化附件表"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS attachment (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entity_type TEXT NOT NULL,
            entity_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            original_name TEXT NOT NULL,
            file_size INTEGER,
            file_type TEXT,
            uploader_id INTEGER,
            uploader_name TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

@app.route('/api/attachments', methods=['GET'])
def get_attachments():
    """获取附件列表"""
    entity_type = request.args.get('entity_type', '')
    entity_id = request.args.get('entity_id', '')
    
    conn = get_db()
    cursor = conn.cursor()
    
    if entity_type and entity_id:
        cursor.execute('''
            SELECT * FROM attachment 
            WHERE entity_type = ? AND entity_id = ?
            ORDER BY created_at DESC
        ''', (entity_type, int(entity_id)))
        attachments = [dict(row) for row in cursor.fetchall()]
    else:
        attachments = []
    
    conn.close()
    return jsonify({'success': True, 'attachments': attachments})

@app.route('/api/attachments/upload', methods=['POST'])
def upload_attachment():
    """上传附件"""
    entity_type = request.form.get('entity_type')
    entity_id = request.form.get('entity_id')
    uploader_name = request.form.get('uploader_name', '匿名')
    
    if not entity_type or not entity_id:
        return jsonify({'success': False, 'error': '缺少关联对象信息'}), 400
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未选择文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '未选择文件'}), 400
    
    # 保存文件
    import uuid
    ext = os.path.splitext(file.filename)[1]
    filename = f"{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    file.save(filepath)
    
    # 记录到数据库
    file_size = os.path.getsize(filepath)
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO attachment (entity_type, entity_id, filename, original_name, file_size, file_type, uploader_name)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (entity_type, int(entity_id), filename, file.filename, file_size, ext, uploader_name))
    attachment_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    # 添加日志
    add_log(entity_type, int(entity_id), 'attachment_added', None, file.filename, None, uploader_name)
    
    return jsonify({
        'success': True,
        'attachment': {
            'id': attachment_id,
            'filename': filename,
            'original_name': file.filename,
            'file_size': file_size
        }
    })

@app.route('/api/attachments/<int:attachment_id>/download', methods=['GET'])
def download_attachment(attachment_id):
    """下载附件"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM attachment WHERE id = ?', (attachment_id,))
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        return jsonify({'success': False, 'error': '附件不存在'}), 404
    
    filepath = os.path.join(UPLOAD_DIR, row['filename'])
    if not os.path.exists(filepath):
        return jsonify({'success': False, 'error': '文件不存在'}), 404
    
    return send_file(filepath, as_attachment=True, download_name=row['original_name'])

@app.route('/api/attachments/<int:attachment_id>', methods=['DELETE'])
def delete_attachment(attachment_id):
    """删除附件"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM attachment WHERE id = ?', (attachment_id,))
    row = cursor.fetchone()
    
    if row:
        # 删除文件
        filepath = os.path.join(UPLOAD_DIR, row['filename'])
        if os.path.exists(filepath):
            os.remove(filepath)
        
        # 删除记录
        cursor.execute('DELETE FROM attachment WHERE id = ?', (attachment_id,))
        conn.commit()
    
    conn.close()
    return jsonify({'success': True, 'deleted': True})

# 初始化附件表
init_attachment_table()

# ========== 通知管理 ==========
def init_notification_table():
    """初始化通知表"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS notification (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                type TEXT NOT NULL DEFAULT 'info',
                title TEXT NOT NULL,
                content TEXT,
                entity_type TEXT,
                entity_id INTEGER,
                is_read INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.commit()

init_notification_table()

def create_notification(user_id, ntype, title, content=None, entity_type=None, entity_id=None):
    """创建通知"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO notification (user_id, type, title, content, entity_type, entity_id)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (user_id, ntype, title, content, entity_type, entity_id))
        conn.commit()
        return cursor.lastrowid

@app.route('/api/notifications', methods=['GET'])
@check_auth
def get_notifications():
    """获取当前用户的通知"""
    user = request.current_user
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM notification WHERE user_id = ? ORDER BY created_at DESC LIMIT 50', (user.get('person_id'),))
        notifications = [dict(row) for row in cursor.fetchall()]
        unread = sum(1 for n in notifications if not n['is_read'])
    return jsonify({'success': True, 'notifications': notifications, 'unread_count': unread})

@app.route('/api/notifications/<int:nid>/read', methods=['POST'])
@check_auth
def mark_notification_read(nid):
    """标记通知为已读"""
    user = request.current_user
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE notification SET is_read = 1 WHERE id = ? AND user_id = ?', (nid, user.get('person_id')))
        conn.commit()
    return jsonify({'success': True})

@app.route('/api/notifications/read-all', methods=['POST'])
@check_auth
def mark_all_notifications_read():
    """标记所有通知为已读"""
    user = request.current_user
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE notification SET is_read = 1 WHERE user_id = ? AND is_read = 0', (user.get('person_id'),))
        conn.commit()
    return jsonify({'success': True, 'updated': cursor.rowcount})

@app.route('/api/notifications/<int:nid>', methods=['DELETE'])
@check_auth
def delete_notification(nid):
    """删除通知"""
    user = request.current_user
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM notification WHERE id = ? AND user_id = ?', (nid, user.get('person_id')))
        conn.commit()
    return jsonify({'success': True})

# ========== 初始化新表 ==========

init_log_table()
init_comment_table()
init_template_table()

# ========== 健康检查 ==========

@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({'status': 'ok', 'service': 'pm-system', 'port': 5236})

# ========== 数据导入导出 ==========
import csv
from openpyxl import Workbook, load_workbook

@app.route('/api/export/tasks', methods=['GET'])
@check_auth
def export_tasks():
    """导出任务为Excel"""
    user = request.current_user
    conn = get_db()
    cursor = conn.cursor()
    
    if user.get('role') == 'admin':
        cursor.execute('''SELECT t.id, t.name, p.name as project, ph.name as phase, 
            per.name as assignee, t.priority, t.status, t.progress, t.due_date, t.description
            FROM task t LEFT JOIN project p ON t.project_id=p.id 
            LEFT JOIN phase ph ON t.phase_id=ph.id
            LEFT JOIN person per ON t.assignee_id=per.id
            ORDER BY t.id''')
    else:
        project_ids = get_user_project_ids(user.get('person_id'))
        if not project_ids:
            conn.close()
            return jsonify({'success': True, 'message': '无数据'})
        placeholders = ','.join(['?'] * len(project_ids))
        cursor.execute(f'''SELECT t.id, t.name, p.name as project, ph.name as phase,
            per.name as assignee, t.priority, t.status, t.progress, t.due_date, t.description
            FROM task t LEFT JOIN project p ON t.project_id=p.id
            LEFT JOIN phase ph ON t.phase_id=ph.id
            LEFT JOIN person per ON t.assignee_id=per.id
            WHERE t.project_id IN ({placeholders})
            ORDER BY t.id''', project_ids)
    
    tasks = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = '任务列表'
    headers = ['ID', '任务名称', '项目', '阶段', '负责人', '优先级', '状态', '进度%', '截止日期', '描述']
    ws.append(headers)
    
    status_map = {'pending': '待处理', 'in_progress': '进行中', 'completed': '已完成', 'cancelled': '已中止'}
    priority_map = {'high': '高', 'medium': '中', 'low': '低'}
    
    for t in tasks:
        ws.append([t['id'], t['name'], t['project'] or '', t['phase'] or '', t['assignee'] or '',
                   priority_map.get(t['priority'], t['priority']), status_map.get(t['status'], t['status']),
                   t['progress'] or 0, t['due_date'] or '', t['description'] or ''])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, as_attachment=True, download_name='tasks_export.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/export/persons', methods=['GET'])
@check_auth
@require_admin
def export_persons():
    """导出人员为Excel"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id, name, line, position, phone, email, status FROM person ORDER BY id')
    persons = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = '人员列表'
    ws.append(['ID', '姓名', '条线', '职位', '电话', '邮箱', '状态'])
    for p in persons:
        ws.append([p['id'], p['name'], p['line'] or '', p['position'] or '', p['phone'] or '', p['email'] or '', p['status'] or ''])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='persons_export.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/import/tasks', methods=['POST'])
@check_auth
@require_admin
def import_tasks():
    """从Excel/CSV导入任务"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '请上传文件'}), 400
    
    file = request.files['file']
    if not file.filename:
        return jsonify({'success': False, 'error': '文件名为空'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    imported = 0
    errors = []
    
    try:
        if file.filename.endswith('.xlsx'):
            wb = load_workbook(file)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
        elif file.filename.endswith('.csv'):
            content = file.read().decode('utf-8')
            rows = list(csv.reader(content.splitlines()))
            rows = rows[1:]  # 跳过表头
        else:
            return jsonify({'success': False, 'error': '仅支持xlsx和csv格式'}), 400
        
        for i, row in enumerate(rows, 2):
            try:
                if not row or not row[1]: continue  # 跳过空行
                name = str(row[1]).strip()
                if not name: continue
                
                # 查找项目
                project_id = None
                if len(row) > 2 and row[2]:
                    cursor.execute('SELECT id FROM project WHERE name = ?', (str(row[2]).strip(),))
                    proj = cursor.fetchone()
                    if proj: project_id = proj['id']
                
                # 查找负责人
                assignee_id = None
                if len(row) > 4 and row[4]:
                    cursor.execute('SELECT id FROM person WHERE name = ?', (str(row[4]).strip(),))
                    person = cursor.fetchone()
                    if person: assignee_id = person['id']
                
                # 优先级映射
                priority = 'medium'
                if len(row) > 5 and row[5]:
                    p = str(row[5]).strip()
                    if p in ('高', 'high', 'H'): priority = 'high'
                    elif p in ('低', 'low', 'L'): priority = 'low'
                
                # 状态映射
                status = 'pending'
                if len(row) > 6 and row[6]:
                    s = str(row[6]).strip()
                    if s in ('进行中', 'in_progress'): status = 'in_progress'
                    elif s in ('已完成', 'completed'): status = 'completed'
                
                progress = 0
                if len(row) > 7 and row[7]:
                    try: progress = int(float(row[7]))
                    except: pass
                
                due_date = None
                if len(row) > 8 and row[8]:
                    due_date = str(row[8]).strip() or None
                
                cursor.execute('''INSERT INTO task (project_id, name, priority, status, progress, due_date, assignee_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?)''', (project_id, name, priority, status, progress, due_date, assignee_id))
                imported += 1
            except Exception as e:
                errors.append(f'第{i}行: {str(e)}')
        
        conn.commit()
    finally:
        conn.close()
    
    return jsonify({'success': True, 'imported': imported, 'errors': errors})

# ========== PDF导出 ==========

import io
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.colors import HexColor, black, white, gray
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# 注册中文字体
FONT_NAME = 'Helvetica'
try:
    # 尝试加载Noto CJK字体（使用subfontIndex）
    pdfmetrics.registerFont(TTFont('NotoSansCJK', '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc', subfontIndex=0))
    FONT_NAME = 'NotoSansCJK'
    print(f"PDF字体加载成功: {FONT_NAME}")
except Exception as e:
    print(f"Noto字体加载失败: {e}")
    try:
        # 后备：使用arphic uming字体
        pdfmetrics.registerFont(TTFont('ARPLUMing', '/usr/share/fonts/truetype/arphic/uming.ttc', subfontIndex=0))
        FONT_NAME = 'ARPLUMing'
        print(f"PDF字体加载成功(后备): {FONT_NAME}")
    except Exception as e2:
        print(f"字体注册全部失败: {e2}")
        FONT_NAME = 'Helvetica'

# PDF风格样式定义（V3版本 - 专业美观）
PDF_STYLES = {
    1: '极简白底表',
    2: '企业蓝报告',
    3: '进度追踪表',
    4: '任务清单式',
    5: '时间轴式',
}

def create_pdf_style1(tasks, project_name, styles, landscape=False):
    """风格1：极简黑白表 - 纯黑白配色，自适应行高显示完整内容"""
    story = []
    status_map = {'pending': '待处理', 'in_progress': '进行中', 'completed': '已完成', 'cancelled': '已中止'}
    
    # 主标题 - 项目名称（醒目）
    title_style = ParagraphStyle('Title', fontName=FONT_NAME, fontSize=18, textColor=black, alignment=1, spaceAfter=6, leading=22)
    story.append(Paragraph(f"<b>{project_name or '任务清单'}</b>", title_style))
    
    # 分隔线
    story.append(Spacer(1, 4))
    
    date_style = ParagraphStyle('Date', fontName=FONT_NAME, fontSize=9, textColor=black, alignment=1, spaceAfter=10)
    story.append(Paragraph(datetime.now().strftime('%Y年%m月%d日'), date_style))
    
    total = len(tasks)
    completed = sum(1 for t in tasks if t.get('status') == 'completed')
    pending = sum(1 for t in tasks if t.get('status') == 'pending')
    in_progress = sum(1 for t in tasks if t.get('status') == 'in_progress')
    
    stats_style = ParagraphStyle('Stats', fontName=FONT_NAME, fontSize=9, textColor=black, alignment=1, spaceAfter=15)
    story.append(Paragraph(f"共 {total} 项 | 完成 {completed} | 进行中 {in_progress} | 待办 {pending}", stats_style))
    
    # 列宽：自适应行高模式，完成情况列足够宽以自动换行
    # 纵向A4约470pt可用宽度，横向约800pt
    if landscape:
        # 横向：完成任务列宽大，可自动换行显示完整内容
        col_widths = [25, 130, 50, 30, 350, 45, 70]
    else:
        # 纵向：压缩其他列，完成情况列足够宽换行
        col_widths = [20, 70, 35, 25, 230, 40, 60]
    
    header_style = ParagraphStyle('Header', fontName=FONT_NAME, fontSize=9, textColor=white, alignment=1)
    header_row = [
        Paragraph('#', header_style),
        Paragraph('任务名称', header_style),
        Paragraph('负责人', header_style),
        Paragraph('进度', header_style),
        Paragraph('完成情况', header_style),
        Paragraph('状态', header_style),
        Paragraph('截止日期', header_style)
    ]
    table_data = [header_row]
    
    # 内容单元格样式 - leading控制换行间距
    cell_style = ParagraphStyle('Cell', fontName=FONT_NAME, fontSize=8, textColor=black, leading=12)
    
    for i, t in enumerate(tasks, 1):
        task_name = (t.get('name', '') or '未命名')
        owner = (t.get('assignee_name') or '-')[:8]
        progress = f"{t.get('progress', 0)}%"
        # 完成情况：优先显示notes（进度备注），其次description（任务描述），合并显示
        # 不截断，让Paragraph自动换行
        notes = (t.get('notes') or '').strip()
        desc_raw = (t.get('description') or '').strip()
        if notes and desc_raw:
            completion = f"{notes} | {desc_raw}"
        elif notes:
            completion = notes
        else:
            completion = desc_raw
        # 不截断，完整显示
        status = status_map.get(t.get('status'), '未知')
        due = (t.get('due_date') or '—')[:10]
        
        row = [
            Paragraph(str(i), cell_style),
            Paragraph(task_name, cell_style),  # 任务名称完整显示
            Paragraph(owner, cell_style),
            Paragraph(progress, cell_style),
            Paragraph(completion if completion else '—', cell_style),  # 完成情况完整显示，自动换行
            Paragraph(status, cell_style),
            Paragraph(due, cell_style)
        ]
        table_data.append(row)
    
    task_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    task_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), FONT_NAME),
        ('BACKGROUND', (0,0), (-1,0), black),
        ('TEXTCOLOR', (0,0), (-1,0), white),
        ('TEXTCOLOR', (0,1), (-1,-1), black),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (1,1), (1,-1), 'LEFT'),
        ('ALIGN', (4,1), (4,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOX', (0,0), (-1,-1), 1.5, black),        # 外框加粗
        ('INNERGRID', (0,0), (-1,-1), 0.8, black),  # 内格加粗
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (1,1), (1,-1), 4),
        ('LEFTPADDING', (4,1), (4,-1), 4),
    ]))
    story.append(task_table)
    
    # 任务统计
    footer_style = ParagraphStyle('Footer', fontName=FONT_NAME, fontSize=8, textColor=black, alignment=1, spaceBefore=15)
    story.append(Paragraph(f"— 共 {total} 项任务 —", footer_style))
    
    # 落款：日期 + 部门名称
    story.append(Spacer(1, 25))
    signature_style = ParagraphStyle('Signature', fontName=FONT_NAME, fontSize=10, textColor=black, alignment=2)
    story.append(Paragraph(datetime.now().strftime('%Y年%m月%d日'), signature_style))
    story.append(Spacer(1, 4))
    dept_style = ParagraphStyle('Dept', fontName=FONT_NAME, fontSize=10, textColor=black, alignment=2)
    story.append(Paragraph("智能与数字化中心·数据智能", dept_style))
    
    return story


def create_pdf_style2(tasks, project_name, styles, landscape=False):
    """风格2：企业蓝报告 - 带完成情况列"""
    story = []
    status_map = {'pending': '待处理', 'in_progress': '进行中', 'completed': '已完成', 'cancelled': '已中止'}
    
    # 主标题区 - 项目名称（蓝色背景框，字号加大）
    width_total = 595 if landscape else 440
    title_data = [[Paragraph(f"<b>{project_name or '项目任务报告'}</b>", 
        ParagraphStyle('T', fontName=FONT_NAME, fontSize=18, textColor=white, alignment=1))]]
    title_table = Table(title_data, colWidths=[width_total])
    title_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,0), HexColor('#1565C0')),
        ('TOPPADDING', (0,0), (0,0), 14),
        ('BOTTOMPADDING', (0,0), (0,0), 14),
    ]))
    story.append(title_table)
    story.append(Spacer(1, 10))
    
    total = len(tasks)
    completed = sum(1 for t in tasks if t.get('status') == 'completed')
    in_progress = sum(1 for t in tasks if t.get('status') == 'in_progress')
    pending = sum(1 for t in tasks if t.get('status') == 'pending')
    avg_progress = sum(t.get('progress', 0) for t in tasks) / total if total > 0 else 0
    
    # 统计卡片
    stat_w = width_total / 5
    stats_data = [
        [Paragraph(f"<b>{total}</b>", ParagraphStyle('SN', fontName=FONT_NAME, fontSize=14, textColor=HexColor('#1565C0'), alignment=1)),
         Paragraph(f"<b>{completed}</b>", ParagraphStyle('SN', fontName=FONT_NAME, fontSize=14, textColor=HexColor('#2E7D32'), alignment=1)),
         Paragraph(f"<b>{in_progress}</b>", ParagraphStyle('SN', fontName=FONT_NAME, fontSize=14, textColor=HexColor('#F57C00'), alignment=1)),
         Paragraph(f"<b>{pending}</b>", ParagraphStyle('SN', fontName=FONT_NAME, fontSize=14, textColor=gray, alignment=1)),
         Paragraph(f"<b>{avg_progress:.0f}%</b>", ParagraphStyle('SN', fontName=FONT_NAME, fontSize=14, textColor=HexColor('#1565C0'), alignment=1))],
        [Paragraph('总任务', ParagraphStyle('SL', fontName=FONT_NAME, fontSize=8, textColor=gray, alignment=1)),
         Paragraph('已完成', ParagraphStyle('SL', fontName=FONT_NAME, fontSize=8, textColor=gray, alignment=1)),
         Paragraph('进行中', ParagraphStyle('SL', fontName=FONT_NAME, fontSize=8, textColor=gray, alignment=1)),
         Paragraph('待处理', ParagraphStyle('SL', fontName=FONT_NAME, fontSize=8, textColor=gray, alignment=1)),
         Paragraph('平均进度', ParagraphStyle('SL', fontName=FONT_NAME, fontSize=8, textColor=gray, alignment=1))]
    ]
    stats_table = Table(stats_data, colWidths=[stat_w]*5)
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), HexColor('#E3F2FD')),
        ('TOPPADDING', (0,0), (0,0), 10),
        ('BOTTOMPADDING', (0,0), (0,0), 4),
        ('TOPPADDING', (0,1), (0,1), 4),
        ('BOTTOMPADDING', (0,1), (0,1), 8),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 15))
    
    date_style = ParagraphStyle('D', fontName=FONT_NAME, fontSize=9, textColor=gray, alignment=2)
    story.append(Paragraph(f"报告日期：{datetime.now().strftime('%Y-%m-%d')}", date_style))
    story.append(Spacer(1, 10))
    
    # 表格列宽：自适应行高，内容自动换行完整显示
    if landscape:
        col_widths = [25, 120, 50, 30, 350, 45, 80]
    else:
        col_widths = [20, 65, 35, 25, 220, 40, 75]
    
    header_style = ParagraphStyle('H', fontName=FONT_NAME, fontSize=9, textColor=white, alignment=1)
    header_row = [
        Paragraph('#', header_style),
        Paragraph('任务名称', header_style),
        Paragraph('负责人', header_style),
        Paragraph('进度', header_style),
        Paragraph('完成情况', header_style),
        Paragraph('状态', header_style),
        Paragraph('截止日期', header_style)
    ]
    table_data = [header_row]
    
    cell_style = ParagraphStyle('C', fontName=FONT_NAME, fontSize=8, textColor=HexColor('#333'), leading=12)
    
    for i, t in enumerate(tasks, 1):
        task_name = (t.get('name', '') or '未命名')
        owner = (t.get('assignee_name') or '-')[:8]
        progress = f"{t.get('progress', 0)}%"
        # 完成情况：不截断，完整显示自动换行
        notes = (t.get('notes') or '').strip()
        desc_raw = (t.get('description') or '').strip()
        if notes and desc_raw:
            completion = f"{notes} | {desc_raw}"
        elif notes:
            completion = notes
        else:
            completion = desc_raw
        # 不截断
        status = status_map.get(t.get('status'), '未知')
        due = (t.get('due_date') or '—')[:10]
        
        row = [
            Paragraph(str(i), cell_style),
            Paragraph(task_name, cell_style),
            Paragraph(owner, cell_style),
            Paragraph(progress, cell_style),
            Paragraph(completion if completion else '—', cell_style),
            Paragraph(status, cell_style),
            Paragraph(due, cell_style)
        ]
        table_data.append(row)
    
    task_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    task_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), FONT_NAME),
        ('BACKGROUND', (0,0), (-1,0), HexColor('#1565C0')),
        ('TEXTCOLOR', (0,0), (-1,0), white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (1,1), (1,-1), 'LEFT'),
        ('ALIGN', (4,1), (4,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, HexColor('#90CAF9')),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (1,1), (1,-1), 4),
        ('LEFTPADDING', (4,1), (4,-1), 4),
    ]))
    story.append(task_table)
    
    # 落款：日期 + 部门名称
    story.append(Spacer(1, 25))
    signature_style = ParagraphStyle('Signature', fontName=FONT_NAME, fontSize=10, textColor=HexColor('#333'), alignment=2)
    story.append(Paragraph(datetime.now().strftime('%Y年%m月%d日'), signature_style))
    story.append(Spacer(1, 4))
    dept_style = ParagraphStyle('Dept', fontName=FONT_NAME, fontSize=10, textColor=HexColor('#1565C0'), alignment=2)
    story.append(Paragraph("智能与数字化中心·数据智能", dept_style))
    
    return story


def create_pdf_style3(tasks, project_name, styles, landscape=False):
    """风格3：进度追踪表 - 进度条可视化，含完成情况"""
    story = []
    status_map = {'pending': '待处理', 'in_progress': '进行中', 'completed': '已完成', 'cancelled': '已中止'}
    
    # 主标题 - 项目名称（醒目）
    title_style = ParagraphStyle('T', fontName=FONT_NAME, fontSize=18, textColor=HexColor('#333'), spaceAfter=6, leading=22)
    story.append(Paragraph(f"<b>📊 {project_name or '项目进度追踪'}</b>", title_style))
    
    story.append(Spacer(1, 4))
    
    date_style = ParagraphStyle('D', fontName=FONT_NAME, fontSize=9, textColor=gray, spaceAfter=12)
    story.append(Paragraph(datetime.now().strftime('%Y-%m-%d %H:%M'), date_style))
    
    total = len(tasks)
    completed = sum(1 for t in tasks if t.get('status') == 'completed')
    avg_progress = sum(t.get('progress', 0) for t in tasks) / total if total > 0 else 0
    
    summary_style = ParagraphStyle('S', fontName=FONT_NAME, fontSize=10, textColor=HexColor('#1565C0'), spaceAfter=12)
    story.append(Paragraph(f"总体进度：{avg_progress:.1f}%  |  已完成：{completed}/{total}", summary_style))
    
    if landscape:
        col_widths = [25, 120, 50, 60, 350, 100]
    else:
        col_widths = [20, 80, 40, 50, 230, 70]
    
    header_style = ParagraphStyle('H', fontName=FONT_NAME, fontSize=9, textColor=HexColor('#333'))
    header_row = [
        Paragraph('#', header_style),
        Paragraph('任务名称', header_style),
        Paragraph('负责人', header_style),
        Paragraph('进度', header_style),
        Paragraph('完成情况', header_style),
        Paragraph('截止', header_style)
    ]
    table_data = [header_row]
    
    cell_style = ParagraphStyle('C', fontName=FONT_NAME, fontSize=8, textColor=HexColor('#333'), leading=12)
    progress_style = ParagraphStyle('P', fontName=FONT_NAME, fontSize=7, textColor=HexColor('#1565C0'))
    
    for i, t in enumerate(tasks, 1):
        task_name = (t.get('name', '') or '未命名')
        owner = (t.get('assignee_name') or '-')[:8]
        progress_val = t.get('progress', 0)
        # 完成情况：不截断，完整显示自动换行
        notes = (t.get('notes') or '').strip()
        desc_raw = (t.get('description') or '').strip()
        if notes and desc_raw:
            completion = f"{notes} | {desc_raw}"
        elif notes:
            completion = notes
        else:
            completion = desc_raw
        # 不截断
        due = (t.get('due_date') or '—')[:10]
        
        progress_text = f"{'█' * (progress_val // 10)}{'░' * (10 - progress_val // 10)} {progress_val}%"
        
        row = [
            Paragraph(str(i), cell_style),
            Paragraph(task_name, cell_style),
            Paragraph(owner, cell_style),
            Paragraph(progress_text, progress_style),
            Paragraph(completion if completion else '—', cell_style),
            Paragraph(due, cell_style)
        ]
        table_data.append(row)
    
    task_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    task_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), FONT_NAME),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('ALIGN', (3,0), (3,-1), 'LEFT'),
        ('ALIGN', (4,1), (4,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LINEBELOW', (0,0), (-1,0), 1, HexColor('#1565C0')),
        ('LINEBELOW', (0,1), (-1,-1), 0.25, HexColor('#E0E0E0')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (1,1), (1,-1), 4),
        ('LEFTPADDING', (4,1), (4,-1), 4),
    ]))
    story.append(task_table)
    
    status_counts = {}
    for t in tasks:
        s = status_map.get(t.get('status'), '未知')
        status_counts[s] = status_counts.get(s, 0) + 1
    
    footer_style = ParagraphStyle('F', fontName=FONT_NAME, fontSize=8, textColor=gray, alignment=1, spaceBefore=12)
    status_str = ' | '.join([f"{k}: {v}" for k, v in status_counts.items()])
    story.append(Paragraph(status_str, footer_style))
    
    # 落款：日期 + 部门名称
    story.append(Spacer(1, 25))
    signature_style = ParagraphStyle('Sig', fontName=FONT_NAME, fontSize=10, textColor=HexColor('#333'), alignment=2)
    story.append(Paragraph(datetime.now().strftime('%Y年%m月%d日'), signature_style))
    story.append(Spacer(1, 4))
    dept_style = ParagraphStyle('Dept', fontName=FONT_NAME, fontSize=10, textColor=HexColor('#1565C0'), alignment=2)
    story.append(Paragraph("智能与数字化中心·数据智能", dept_style))
    
    return story


def create_pdf_style4(tasks, project_name, styles, landscape=False):
    """风格4：任务清单式 - 含完成情况详情"""
    story = []
    status_map = {'pending': '待处理', 'in_progress': '进行中', 'completed': '已完成', 'cancelled': '已中止'}
    
    # 主标题 - 项目名称（醒目）
    title_style = ParagraphStyle('T', fontName=FONT_NAME, fontSize=18, textColor=black, spaceAfter=6, leading=22)
    story.append(Paragraph(f"<b>☑ {project_name or '任务清单'}</b>", title_style))
    
    story.append(Spacer(1, 4))
    
    date_style = ParagraphStyle('D', fontName=FONT_NAME, fontSize=9, textColor=black, spaceAfter=10)
    story.append(Paragraph(datetime.now().strftime('%Y年%m月%d日'), date_style))
    
    for i, t in enumerate(tasks, 1):
        task_name = (t.get('name', '') or '未命名')
        owner = (t.get('assignee_name') or '-')
        status = status_map.get(t.get('status'), '未知')
        progress = t.get('progress', 0)
        due = t.get('due_date') or '未设定'
        # 完成情况：不截断，完整显示
        notes = (t.get('notes') or '').strip()
        desc_raw = (t.get('description') or '').strip()
        if notes and desc_raw:
            completion = f"{notes} | {desc_raw}"
        elif notes:
            completion = notes
        else:
            completion = desc_raw
        
        status_mark = {'待处理': '○', '进行中': '◐', '已完成': '●', '已中止': '✕', '未知': '?'}
        mark = status_mark.get(status, '?')
        
        item_style = ParagraphStyle('Item', fontName=FONT_NAME, fontSize=9, textColor=black, leading=14, spaceBefore=4, spaceAfter=4)
        item_text = f"{mark} <b>{task_name}</b>"
        story.append(Paragraph(item_text, item_style))
        
        detail_style = ParagraphStyle('Detail', fontName=FONT_NAME, fontSize=8, textColor=black, leading=12, leftIndent=20)
        detail_text = f"负责人：{owner[:10]}  |  进度：{progress}%  |  截止：{due[:12]}"
        story.append(Paragraph(detail_text, detail_style))
        
        # 完成情况单独一行显示（可能很长）
        if completion:
            comp_style = ParagraphStyle('Comp', fontName=FONT_NAME, fontSize=8, textColor=gray, leading=12, leftIndent=20)
            story.append(Paragraph(f"完成情况：{completion}", comp_style))
        
        story.append(Spacer(1, 4))
    
    total = len(tasks)
    completed = sum(1 for t in tasks if t.get('status') == 'completed')
    footer_style = ParagraphStyle('F', fontName=FONT_NAME, fontSize=9, textColor=black, alignment=1, spaceBefore=10)
    story.append(Paragraph(f"— 共 {total} 项，已完成 {completed} 项 —", footer_style))
    
    # 落款：日期 + 部门名称
    story.append(Spacer(1, 25))
    signature_style = ParagraphStyle('Sig', fontName=FONT_NAME, fontSize=10, textColor=black, alignment=2)
    story.append(Paragraph(datetime.now().strftime('%Y年%m月%d日'), signature_style))
    story.append(Spacer(1, 4))
    dept_style = ParagraphStyle('Dept', fontName=FONT_NAME, fontSize=10, textColor=black, alignment=2)
    story.append(Paragraph("智能与数字化中心·数据智能", dept_style))
    
    return story


def create_pdf_style5(tasks, project_name, styles, landscape=False):
    """风格5：时间轴式 - 含完成情况"""
    story = []
    status_map = {'pending': '待处理', 'in_progress': '进行中', 'completed': '已完成', 'cancelled': '已中止'}
    
    # 主标题 - 项目名称（醒目）
    title_style = ParagraphStyle('T', fontName=FONT_NAME, fontSize=18, textColor=HexColor('#1565C0'), spaceAfter=6, leading=22)
    story.append(Paragraph(f"<b>📅 {project_name or '任务时间轴'}</b>", title_style))
    
    story.append(Spacer(1, 4))
    
    sorted_tasks = sorted(tasks, key=lambda x: x.get('due_date') or '9999-99-99')
    
    prev_date = None
    for i, t in enumerate(sorted_tasks, 1):
        task_name = (t.get('name', '') or '未命名')
        owner = (t.get('assignee_name') or '-')
        status = status_map.get(t.get('status'), '未知')
        progress = t.get('progress', 0)
        due = t.get('due_date') or '未设定'
        # 完成情况：不截断，完整显示
        notes = (t.get('notes') or '').strip()
        desc_raw = (t.get('description') or '').strip()
        if notes and desc_raw:
            completion = f"{notes} | {desc_raw}"
        elif notes:
            completion = notes
        else:
            completion = desc_raw
        
        if due != prev_date and due != '未设定':
            date_header_style = ParagraphStyle('DH', fontName=FONT_NAME, fontSize=10, textColor=HexColor('#1565C0'), spaceBefore=10, spaceAfter=6)
            story.append(Paragraph(f"━━━ {due} ━━━", date_header_style))
            prev_date = due
        
        node_style = ParagraphStyle('Node', fontName=FONT_NAME, fontSize=9, textColor=HexColor('#333'), leading=13, leftIndent=15)
        node_text = f"▶ <b>{task_name}</b>"
        story.append(Paragraph(node_text, node_style))
        
        detail_style = ParagraphStyle('Det', fontName=FONT_NAME, fontSize=8, textColor=gray, leading=11, leftIndent=25)
        detail_text = f"{owner[:10]} | {status} | 进度{progress}%"
        story.append(Paragraph(detail_text, detail_style))
        
        # 完成情况单独一行显示（可能很长）
        if completion:
            comp_style = ParagraphStyle('Comp2', fontName=FONT_NAME, fontSize=8, textColor=gray, leading=11, leftIndent=25)
            story.append(Paragraph(f"完成：{completion}", comp_style))
        
        story.append(Spacer(1, 3))
    
    total = len(tasks)
    footer_style = ParagraphStyle('F', fontName=FONT_NAME, fontSize=8, textColor=gray, alignment=1, spaceBefore=12)
    story.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}  |  共 {total} 项任务", footer_style))
    
    # 落款：日期 + 部门名称
    story.append(Spacer(1, 25))
    signature_style = ParagraphStyle('Sig', fontName=FONT_NAME, fontSize=10, textColor=HexColor('#333'), alignment=2)
    story.append(Paragraph(datetime.now().strftime('%Y年%m月%d日'), signature_style))
    story.append(Spacer(1, 4))
    dept_style = ParagraphStyle('Dept', fontName=FONT_NAME, fontSize=10, textColor=HexColor('#1565C0'), alignment=2)
    story.append(Paragraph("智能与数字化中心·数据智能", dept_style))
    
    return story


def create_pdf(tasks, style, project_name, landscape=False):
    """创建PDF文档，支持横向模式"""
    buffer = io.BytesIO()
    
    from reportlab.lib.pagesizes import landscape as landscape_mode
    pagesize = landscape_mode(A4) if landscape else A4
    
    doc = SimpleDocTemplate(buffer, pagesize=pagesize, leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    
    styles = getSampleStyleSheet()
    
    style = min(max(style, 1), 5)
    if style == 1:
        story = create_pdf_style1(tasks, project_name, styles, landscape)
    elif style == 2:
        story = create_pdf_style2(tasks, project_name, styles, landscape)
    elif style == 3:
        story = create_pdf_style3(tasks, project_name, styles, landscape)
    elif style == 4:
        story = create_pdf_style4(tasks, project_name, styles, landscape)
    else:
        story = create_pdf_style5(tasks, project_name, styles, landscape)
    
    doc.build(story)
    buffer.seek(0)
    return buffer


@app.route('/api/tasks/export-pdf', methods=['POST'])
def export_tasks_pdf():
    """导出任务PDF，支持状态筛选和横向模式"""
    data = request.json
    tasks = data.get('tasks', [])
    style = data.get('style', 1)
    project_name = data.get('project_name', '全部项目')
    landscape = data.get('landscape', False)
    status_filter = data.get('status', 'all')
    
    # 根据状态筛选任务
    if status_filter != 'all':
        tasks = [t for t in tasks if t.get('status') == status_filter]
    
    if not tasks:
        status_names = {'all': '全部', 'pending': '待处理', 'in_progress': '进行中', 'completed': '已完成', 'cancelled': '已中止'}
        return jsonify({'error': f'没有{status_names.get(status_filter, status_filter)}状态的任务'}), 400
    
    try:
        pdf_buffer = create_pdf(tasks, style, project_name, landscape)
        
        status_names = {'all': '全部', 'pending': '待处理', 'in_progress': '进行中', 'completed': '已完成', 'cancelled': '已中止'}
        orient = '横向' if landscape else '纵向'
        filename = f"tasks_{status_names.get(status_filter, '全部')}_{orient}_{datetime.now().strftime('%Y%m%d')}.pdf"
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"PDF生成错误: {e}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    print("🚀 项目管理系统启动...")
    print(f"   数据库: {DB_PATH}")
    print(f"   端口: 5236")
    print(f"   API: http://localhost:5236/api")
    app.run(host='0.0.0.0', port=5236, debug=False)